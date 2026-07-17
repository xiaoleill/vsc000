import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "验证计划"

# ── Styles ──────────────────────────────────────────
hdr_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")
hdr_fill   = PatternFill("solid", fgColor="4472C4")
hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font  = Font(name="Arial", size=10)
cell_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))
cat_fill   = PatternFill("solid", fgColor="D6E4F0")  # light blue for group headers

col_widths = {"A": 6, "B": 18, "C": 26, "D": 32, "E": 26, "F": 40, "G": 62, "H": 40, "I": 14, "J": 22}

for col_letter, w in col_widths.items():
    ws.column_dimensions[col_letter].width = w

headers = ["序号", "功能大项", "功能子项", "验证项目", "验证Case名称", "验证目标", "验证方法与条件", "期待结果", "验证结果", "备注"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = thin_border

ws.freeze_panes = "A2"

# ── Data ────────────────────────────────────────────
# Each tuple: (功能大项, 功能子项, 验证项目, Case名称, 验证目标, 验证方法, 期待结果, 备注)
rows = [
    # ===================================================
    # 1. 寄存器访问验证
    # ===================================================
    ("寄存器访问验证", "CFGLOCK/PERLOCK机制", "CFGLOCK默认锁定状态检查", "ermu_reg_test", "确认上电后CFGLOCK默认=0x1(锁定)，受保护寄存器写操作被忽略", "上电后不写CFGLOCK，直接尝试写CTCMP等受保护寄存器，再读取验证", "受保护寄存器写入被忽略，读回值不变；CFG_LOCK=1", "CFGLOCK.PERLOCK检查已在reg_test中实现"),
    ("寄存器访问验证", "CFGLOCK/PERLOCK机制", "CFGLOCK解锁验证(KEY=0xBC)", "ermu_reg_test", "确认写入0xBC到KEY字段后CFG_LOCK=0", "写CFGLOCK.KEY=0xBC，读CFGLOCK验证CFG_LOCK=0", "CFG_LOCK=0，受保护寄存器可正常写入", "已在reg_test中实现"),
    ("寄存器访问验证", "CFGLOCK/PERLOCK机制", "CFGLOCK重锁验证(KEY≠0xBC)", "ermu_cfglock_test", "确认写入非0xBC值到KEY字段使CFG_LOCK重新=1", "先解锁→写KEY=任意非0xBC值→读CFGLOCK", "CFG_LOCK=1，后续受保护寄存器写操作再次被忽略", "待实现"),
    ("寄存器访问验证", "CFGLOCK/PERLOCK机制", "PERLOCK永久锁定验证(KEY=0xFF)", "ermu_cfglock_test", "确认写入0xFF永久锁定后所有写操作被忽略", "先解锁CFGLOCK→写PERLOCK.KEY=0xFF→读PERLOCK→尝试写任意受保护寄存器→尝试写CFGLOCK解锁", "PER_LOCK=1；所有受保护寄存器写操作被忽略；CFGLOCK无法再解锁；仅上电复位可恢复", "待实现"),
    ("寄存器访问验证", "CFGLOCK/PERLOCK机制", "EOyC.CLR和WTzC.STP不受CFGLOCK保护", "ermu_cfglock_test", "确认这两个WO位在CFGLOCK锁定状态下仍可写入", "CFGLOCK锁定状态下，写EOyC.CLR和WTzC.STP，验证操作有效", "CLR和STP操作正常执行，不受CFGLOCK影响", "待实现"),
    ("寄存器访问验证", "CFGLOCK/PERLOCK机制", "PERLOCK优先级高于CFGLOCK", "ermu_cfglock_test", "确认PERLOCK=1时即使CFG_LOCK=0也无法写入", "解锁CFGLOCK→设置PERLOCK→尝试写受保护寄存器", "所有受保护寄存器写操作被忽略，PERLOCK具有最高优先级", "待实现"),

    ("寄存器访问验证", "复位值检查", "全寄存器复位值验证", "ermu_reg_test", "确认所有寄存器上电复位值符合设计规格", "复位释放后读取所有寄存器，逐一比对期望复位值", "EOyS=0x1(EOS=1), CFGLOCK=0x1, 其余寄存器=0x0", "已实现，覆盖288源规模"),
    ("寄存器访问验证", "复位值检查", "ERMU0_EC1EG特殊复位值验证", "ermu_reg_test", "确认EC1EG复位值为0x0006_6666(WDT默认APP_RST)", "若DUT支持，读EC1EG验证复位值", "EC1EG=0x0006_6666 (若DUT实现)", "DUT可能不支持此特性"),

    ("寄存器访问验证", "RW/RO/WO属性", "RW寄存器读写验证", "ermu_reg_test", "确认RW寄存器写后读回正确", "写特定pattern→读回，验证数据一致(考虑保留位掩码)", "写入值与读回值一致(保留位除外)", "已实现"),
    ("寄存器访问验证", "RW/RO/WO属性", "RO寄存器写忽略验证", "ermu_reg_test", "确认ESS等RO寄存器无法写入", "写全1到ESS寄存器→读回验证", "ESS值不变", "已实现"),
    ("寄存器访问验证", "RW/RO/WO属性", "WO寄存器读返回0验证", "ermu_reg_test", "确认PET/ESSC等WO寄存器读回0", "写PET寄存器→读回验证", "读回值为0", "已实现"),
    ("寄存器访问验证", "RW/RO/WO属性", "ERCp保留位读回0验证", "ermu_reg_test", "确认每源3bit配置中的保留位(第4bit)读回0", "写ERCp=全1→读回验证", "高位保留位=0，实际读回0x77777777", "已实现"),
    ("寄存器访问验证", "仅字写访问", "非字写访问行为验证", "ermu_reg_test", "确认仅32bit字写被接受，字节/半字写可能被忽略或报错", "尝试通过字节使能(pstrb≠4'b1111)写寄存器→读回", "非字写操作被忽略或产生pslverr", "APB driver已强制pstrb=4'b1111"),

    ("寄存器访问验证", "写保护机制", "CTE=1时写CTCMP被忽略", "ermu_timer_test", "确认清除定时器使能期间无法修改比较值", "先设CTE=1→写CTCMP新值→读回CTCMP", "CTCMP保持原值不变", "待验证"),
    ("寄存器访问验证", "写保护机制", "TTE=1时写TTCMP被忽略", "ermu_timer_test", "确认翻转定时器使能期间无法修改比较值", "先设TTE=1→写TTCMP新值→读回TTCMP", "TTCMP保持原值不变", "待验证"),
    ("寄存器访问验证", "写保护机制", "WTE=1时写WTzCMP被忽略", "ermu_timer_test", "确认等待定时器使能期间无法修改比较值", "先设WTE=1→写WTzCMP新值→读回WTzCMP", "WTzCMP保持原值不变", "待验证"),
    ("寄存器访问验证", "写保护机制", "WTzCMP=1写入禁止", "ermu_timer_test", "确认禁止将等待定时器比较值设为1", "写WTzCMP=1→读回WTzCMP", "WTzCMP≠1，写入被硬件阻止", "待验证"),
    ("寄存器访问验证", "写保护机制", "CTS=1期间写CLR被忽略", "ermu_timer_test", "确认清除定时器计数期间无法软件清除EOS", "触发错误使CTS=1→写CLR→读EOS", "EOS保持1，CLR操作被忽略", "待验证"),

    # ===================================================
    # 2. 错误源管理
    # ===================================================
    ("错误源管理", "错误源检测与状态记录", "单错误源注入→ESS位设置", "ermu_error_src_test", "确认外部err_src_id信号上升沿使对应ESS位置1", "使能某错误源→通过input agent注入脉冲→读ESSj验证对应位", "对应ESSj.ESSk=1，其余位=0", "已实现26个代表源测试"),
    ("错误源管理", "错误源检测与状态记录", "边界错误源验证(ID=0, ID=283)", "ermu_error_src_test", "确认最小和最大外部错误源ID可正确检测", "分别注入ID=0和ID=283，验证ESS正确记录", "ESS0[0]=1(ID=0), ESS8[27]=1(ID=283)", "已实现边界测试"),
    ("错误源管理", "错误源检测与状态记录", "全288源分类抽样验证", "ermu_error_src_test", "确认各模块类别的错误源均能正确检测", "按类别(WT/SWDT/FLASH/SRAM/CACHE/CAN/DMPU/DMA/CPU等)每类至少1个源注入", "所有注入源对应ESS位正确置位", "已实现"),
    ("错误源管理", "错误源检测与状态记录", "内部WT错误源(ID=284~287)状态记录", "ermu_timer_test", "确认wait timer超时产生的内部错误正确记录到ESS", "配置WT→使能→触发超时→读ESS8", "ESS8对应位(WTO timeout=ID 284→bit28)=1", "待验证"),

    ("错误源管理", "错误源状态清除(ESSC)", "ESSC写1清除对应ESS位", "ermu_error_src_test", "确认ESSC写1可清除对应ESS状态位", "注入错误→ESS=1→写ESSC对应位=1→读ESS", "对应ESS位被清除为0", "已实现"),
    ("错误源管理", "错误源状态清除(ESSC)", "ESSC写0不影响ESS位", "ermu_error_src_test", "确认ESSC写0不会清除ESS状态位", "注入错误→ESS=1→写ESSC对应位=0→读ESS", "ESS位保持1不变", "待实现"),
    ("错误源管理", "错误源状态清除(ESSC)", "全组清除验证(写全1)", "ermu_error_src_test", "确认写全1到ESSC可一次性清除整组32个ESS位", "注入多个错误源→写ESSC=0xFFFFFFFF→读ESS", "整组ESS全部清零", "待实现"),

    ("错误源管理", "伪错误触发(PET)", "PET写1触发伪错误", "ermu_pseudo_err_test", "确认通过PET寄存器可模拟错误源触发", "不操作外部err_src_id，直接写PET对应位=1→读ESS", "对应ESS位置1，行为与真实错误一致", "待实现"),
    ("错误源管理", "伪错误触发(PET)", "PET全288源伪触发验证", "ermu_pseudo_err_test", "确认所有288源均可通过PET触发", "遍历所有PET寄存器组，逐一触发→读ESS验证", "所有源对应ESS位正确置位", "待实现"),
    ("错误源管理", "伪错误触发(PET)", "PET写0不触发错误", "ermu_pseudo_err_test", "确认PET写0不会产生伪错误", "写PET=0→读ESS", "ESS保持全0", "待实现"),

    ("错误源管理", "错误源边沿触发行为", "错误源持续保持时ESS行为", "ermu_error_src_test", "确认错误源信号持续为高时ESS保持=1", "驱动err_src_id保持高→读ESS→过一段时间再读ESS", "ESS保持1，不清零直到写ESSC", "待实现"),
    ("错误源管理", "错误源边沿触发行为", "错误源脉冲宽度要求", "ermu_error_src_test", "确认最小脉冲宽度能被检测到", "注入不同pulse_duration(1/2/5 cycles)→读ESS", "≥1个clk_hrc周期脉冲即可被检测", "待实现"),

    ("错误源管理", "多错误源并发", "同组多错误源并发检测", "ermu_concurrent_test", "确认同一ESS组内多个错误源同时触发能被全部记录", "同时注入同一组内多个错误源→读ESS", "所有注入源的对应ESS位全为1", "待实现"),
    ("错误源管理", "多错误源并发", "跨组多错误源并发检测", "ermu_concurrent_test", "确认跨ESS组的多个错误源同时触发能被全部记录", "同时注入跨组的多个错误源→读所有相关ESS", "各组的对应ESS位均正确置位", "待实现"),

    # ===================================================
    # 3. 错误响应控制
    # ===================================================
    ("错误响应控制", "ERMU_NA (0x0)", "NA响应验证", "ermu_error_rsp_test", "确认配置NA响应时不产生任何中断/复位/输出", "配置某源ERC=0x0→注入错误→检查HPI/LPI/复位请求/EOUT", "无任何中断、复位请求或错误输出变化", "待实现"),
    ("错误响应控制", "ERMU_LPI0 (0x2)", "LPI0中断产生验证", "ermu_error_rsp_test", "确认配置LPI0时错误产生对应低优先级中断", "配置ERC=0x2→注入错误→监测lpi_irq_o[0]", "lpi_irq_o[0]=1", "待实现"),
    ("错误响应控制", "ERMU_LPI1 (0x3)", "LPI1中断产生验证", "ermu_error_rsp_test", "确认配置LPI1时错误产生对应低优先级中断", "配置ERC=0x3→注入错误→监测lpi_irq_o[1]", "lpi_irq_o[1]=1", "待实现"),
    ("错误响应控制", "ERMU_LPI2 (0x4)", "LPI2中断产生验证", "ermu_error_rsp_test", "确认配置LPI2时错误产生对应低优先级中断", "配置ERC=0x4→注入错误→监测lpi_irq_o[2]", "lpi_irq_o[2]=1", "待实现"),
    ("错误响应控制", "ERMU_HPI (0x5)", "HPI中断产生验证", "ermu_hpi_test", "确认配置HPI响应且HPIE使能时产生NMI", "配置ERC=0x5→设EGC.HPIE=01→注入错误→监测hpi_irq_o", "hpi_irq_o对应位=1", "待实现"),
    ("错误响应控制", "ERMU_HPI (0x5)", "HPIE=00时HPI不产生", "ermu_hpi_test", "确认HPIE禁用时即使ERC配HPI也不产生中断", "设EGC.HPIE=00→配置ERC=0x5→注入错误→监测hpi_irq_o", "hpi_irq_o=0", "待实现"),
    ("错误响应控制", "ERMU_HPI (0x5)", "HPI双核路由(HPIE_C0+C1)", "ermu_hpi_test", "确认HPI可同时路由到两个核心", "设EGC.HPIE=11→配置ERC=0x5→注入错误→监测hpi_irq_o", "hpi_irq_o[1:0]=11，两个核均收到NMI", "待实现"),
    ("错误响应控制", "ERMU_APP_RST (0x6)", "应用复位请求验证", "ermu_error_rsp_test", "确认配置APP_RST时产生应用复位请求", "配置ERC=0x6→注入错误→监测arst_req_o", "arst_req_o=1", "待实现"),
    ("错误响应控制", "ERMU_SYS_RST (0x7)", "系统复位请求验证", "ermu_error_rsp_test", "确认配置SYS_RST时产生系统复位请求", "配置ERC=0x7→注入错误→监测srst_req_o", "srst_req_o=1", "待实现"),
    ("错误响应控制", "ERMU_RSV (0x1)", "保留编码无响应验证", "ermu_error_rsp_test", "确认配置RSV时不产生任何意外响应", "配置ERC=0x1→注入错误→监测所有输出", "无任何中断/复位/输出变化", "待实现"),
    ("错误响应控制", "全8种编码遍历", "ExRC编码遍历验证", "ermu_error_rsp_test", "确认所有8种ERC编码行为符合预期", "遍历0x0~0x7写入同一源ERC→注入错误→验证响应", "每种编码的响应行为符合规格", "待实现"),
    ("错误响应控制", "EC1EG默认值", "WDT错误源默认APP_RST验证", "ermu_error_rsp_test", "确认WDT相关错误源上电默认响应为APP_RST", "不配置ERC→直接注入WDT错误→监测arst_req_o", "arst_req_o=1 (若DUT支持EC1EG=0x6666)", "DUT可能不支持"),

    # ===================================================
    # 4. 错误输出通道
    # ===================================================
    ("错误输出通道", "EOS控制", "软件SET置位EOS", "ermu_timer_test", "确认写EOyC.SET=1可使EOS=1", "写EOyC.SET=1→延时→读EOyS", "EOyS.EOS=1", "已实现"),
    ("错误输出通道", "EOS控制", "软件CLR清除EOS(条件满足时)", "ermu_timer_test", "确认当所有unmasked ESS=0且CTS=0时CLR可清除EOS", "确保ESS=0,CTS=0→写CLR=1→读EOyS", "EOyS.EOS=0", "已实现"),
    ("错误输出通道", "EOS控制", "EOS复位值=1验证", "ermu_reg_test", "确认上电后EOS默认为1", "上电后直接读EOyS(不写任何寄存器)", "EOyS.EOS=1", "已实现"),
    ("错误输出通道", "EOS控制", "ESS≠0时CLR不能清除EOS", "ermu_output_chan_test", "确认存在未清除错误时CLR操作无效", "注入错误使ESS≠0→写CLR=1→读EOS", "EOS保持=1", "待实现"),
    ("错误输出通道", "EOS控制", "EOS_SET→CLR→SET循环测试", "ermu_output_chan_test", "确认EOS可多次SET/CLR循环", "SET→验证EOS=1→清除错误→CLR→验证EOS=0→SET→验证EOS=1", "每次SET/CLR后EOS状态正确切换", "待实现"),

    ("错误输出通道", "输出电平", "无错误无Toggle: EOUTyM=H, EOUTyC=L", "ermu_output_chan_test", "确认常态下输出电平正确", "EOS=0,TTE=0→监测eoutm_o/eoutc_o", "eoutm_o=1, eoutc_o=0", "待实现"),
    ("错误输出通道", "输出电平", "有错误无Toggle: EOUTyM=L, EOUTyC=H", "ermu_output_chan_test", "确认错误状态下输出电平翻转", "注入错误使ESS≠0,TTE=0→监测eoutm_o/eoutc_o", "eoutm_o=0, eoutc_o=1", "待实现"),
    ("错误输出通道", "输出电平", "无错误有Toggle: 输出翻转", "ermu_timer_test", "确认Toggle模式输出周期性翻转", "EOS=0,TTE=1,TTCMP配置→监测eoutm_o/eoutc_o", "输出按TTCMP周期翻转，且eoutm_o与eoutc_o互补", "待验证(等timer_test修复)"),
    ("错误输出通道", "输出电平", "有错误有Toggle: 错误优先(输出=L/H)", "ermu_output_chan_test", "确认有错误时即使TTE=1也保持错误输出电平", "先Toggle→注入错误→监测输出", "输出回到错误状态(eoutm_o=0, eoutc_o=1)，不继续翻转", "待实现"),

    ("错误输出通道", "错误输出掩码(OM)", "单源掩码验证", "ermu_error_mask_test", "确认OMk=1时对应源不触发错误输出", "配OMk=1(掩码某源)→注入该源错误→读EOyS", "EOyS.EOS不因该源变化(被掩码)", "待实现"),
    ("错误输出通道", "错误输出掩码(OM)", "OMk=0不掩码验证", "ermu_error_mask_test", "确认OMk=0时对应源正常触发错误输出", "配OMk=0→注入错误→读EOyS", "EOyS.EOS正确反映错误状态", "待实现"),
    ("错误输出通道", "错误输出掩码(OM)", "全掩码(OM=全1)验证", "ermu_error_mask_test", "确认所有源被掩码时任何错误不影响EOS", "配OM0/OM1=全1→注入多个错误→读EOyS", "EOS=0(所有错误被掩码)", "待实现"),
    ("错误输出通道", "错误输出掩码(OM)", "OM受CFGLOCK/PERLOCK保护", "ermu_error_mask_test", "确认掩码寄存器受配置保护", "CFGLOCK锁定→尝试写OM→读回", "OM值不变", "待实现"),

    ("错误输出通道", "4通道独立操作", "各通道独立EOS控制", "ermu_output_chan_test", "确认4个通道的EOS相互独立", "通道0 SET→通道1 CLR→通道2 SET→通道3 CLR→读所有EOyS", "各通道EOS按各自操作正确反映，互不干扰", "待实现"),
    ("错误输出通道", "4通道独立操作", "各通道独立OM掩码", "ermu_output_chan_test", "确认各通道掩码独立配置", "通道0掩码源0→通道1不掩码源0→注入源0→读各通道EOyS", "通道0: EOS=0(掩码), 通道1: EOS=1(未掩码)", "待实现"),
    ("错误输出通道", "4通道独立操作", "各通道独立定时器", "ermu_output_chan_test", "确认各通道clear/toggle timer独立运行", "通道0 CTCMP=100→通道1 CTCMP=200→同时SET→读各自CTCNT", "两通道独立计数，互不影响", "待实现"),

    ("错误输出通道", "多通道并行", "4通道同时错误输出", "ermu_concurrent_test", "确认4通道可同时处理不同错误", "4通道配置不同响应→同时注入4个错误→监测4路输出", "4路输出均正确反映各自配置的响应", "待实现"),

    # ===================================================
    # 5. 定时器功能
    # ===================================================
    ("定时器功能", "Clear Timer (清除定时器)", "基本功能: CTCMP=100超时后自动清除EOS", "ermu_timer_test", "确认clear timer到达阈值后自动清零EOS并停止", "SET EOS→CTE=1, CTCMP=100→等待>100个计数周期→读EOyS", "EOyS.CTS=0(定时器停), EOS=0(已自动清除)", "已实现基本测试"),
    ("定时器功能", "Clear Timer (清除定时器)", "新错误源重启计数器", "ermu_timer_test", "确认计数过程中新错误出现时计数器从1重新开始", "启动clear timer→计数中途注入第二个错误→读CTCNT", "CTCNT从1重新计数(验证CNT变小)", "待实现"),
    ("定时器功能", "Clear Timer (清除定时器)", "CTS=1期间CLR被忽略", "ermu_timer_test", "确认计数期间不能软件清除EOS", "SET+CTE→等CTS=1→写CLR→读EOS", "EOS保持=1，CLR操作被忽略", "待实现"),
    ("定时器功能", "Clear Timer (清除定时器)", "写0到CTE(CTS=0时)停止定时器", "ermu_timer_test", "确认CTE=0且CTS=0时正常停止", "CTE=1→等CTS=0→写CTE=0→验证定时器停止", "CTS保持=0，定时器不再计数", "待实现"),

    ("定时器功能", "Toggle Timer (翻转定时器)", "基本功能: EOS=0时计数到TTCMP翻转输出", "ermu_timer_test", "确认Toggle timer在EOS=0且TTE=1时正常计数", "SET→配TTCMP=50,TTE=1→CLR(EOS→0,开始计数)→等→读TTCNT", "TTCNT>0(正在计数)", "已修复(待验证)"),
    ("定时器功能", "Toggle Timer (翻转定时器)", "使能序列: 配TTCMP→CLR→等EOS=0→TTE=1", "ermu_timer_test", "确认必须按正确序列才能使能Toggle timer", "按手册序列执行→验证TTCNT>0", "Toggle timer正常计数", "已修复流程"),
    ("定时器功能", "Toggle Timer (翻转定时器)", "禁止序列: SET→等EOS=1→TTE=0", "ermu_timer_test", "确认禁止序列正确停止Toggle timer", "Toggle运行中→SET→等EOS=1→TTE=0→读TTCNT和EOyS", "TTCNT停止变化，EOS=1", "待实现"),
    ("定时器功能", "Toggle Timer (翻转定时器)", "输出互补翻转验证", "ermu_timer_test", "确认EOUTyM和EOUTyC互补翻转", "使能Toggle→监测eoutm_o/eoutc_o→比较两信号", "eoutm_o与eoutc_o始终互补", "待实现"),

    ("定时器功能", "Wait Timer (等待定时器)", "LPI触发启动等待定时器", "ermu_timer_test", "确认LPI中断可启动wait timer", "配WTE=1,LPISE=1,CMP=200→注入LPI响应的错误→等超时→读WTzS和ESS", "WTS先=1(计数中), 超时后ESS对应位=1", "已实现基本测试(待验证)"),
    ("定时器功能", "Wait Timer (等待定时器)", "HPI触发启动等待定时器", "ermu_timer_test", "确认HPI中断可启动wait timer", "配WTE=1,HPISE=1,CMP=200→注入HPI响应的错误→等超时→读WTzS和ESS", "WTS先=1(计数中), 超时后ESS对应位=1", "待实现"),
    ("定时器功能", "Wait Timer (等待定时器)", "计数期间新中断不重启", "ermu_timer_test", "确认wait timer计数中收到新中断不影响当前计数", "启动WT→中途注入第二个可启动WT的中断→读WTCNT", "WTCNT不归零，继续从当前值递增", "待实现"),
    ("定时器功能", "Wait Timer (等待定时器)", "ERMU_WTzE内部错误产生验证", "ermu_timer_test", "确认WT超时产生ERMU_WTzE信号可被ESS记录", "WT超时→读ESS8", "WT0_TO(ID=284)→ESS8[28]=1; WT1_TO(ID=285)→ESS8[29]=1", "已增加ESS8检查"),
    ("定时器功能", "Wait Timer (等待定时器)", "软件STP停止等待定时器", "ermu_timer_test", "确认写WTzC.STP=1可停止等待定时器", "WT计数中→写STP=1→读WTzS", "WTS=0，定时器立即停止", "待实现"),
    ("定时器功能", "Wait Timer (等待定时器)", "循环自触发防护验证", "ermu_timer_test", "确认避免因WTzE响应配置与触发源相同导致的死循环", "WT触发=LPI, WTzE响应=HPI→验证不会死循环", "系统不产生死循环或意外复位", "待实现"),

    ("定时器功能", "时钟预分频(CCPS)", "CCPS_PSE使能与禁止", "ermu_prescaler_test", "确认预分频器使能控制正确", "设CCPS.PSE=0→配CTCMP=100→测定时器周期→与PSE=1周期对比", "PSE=0时使用MRC直通时钟(8MHz), PSE=1使用分频后时钟", "待实现"),
    ("定时器功能", "时钟预分频(CCPS)", "PSC[7:0]分频精度验证(各档位)", "ermu_prescaler_test", "确认分频系数公式fcntclk=fMRC/(PSC+1)正确", "遍历PSC=0/1/127/255→测定时器实际周期", "实测周期=125ns×(PSC+1)", "待实现"),
    ("定时器功能", "时钟预分频(CCPS)", "CCPS影响所有3种定时器", "ermu_prescaler_test", "确认CCPS同时影响clear/toggle/wait timer", "设PSC=某值→分别测3种定时器计数频率", "3种定时器计数频率一致，均受CCPS影响", "待实现"),

    ("定时器功能", "Debug行为", "定时器在Debug断点期间不停止", "ermu_timer_test", "确认Core halt时定时器继续计数", "启动定时器→(模拟core halt)→等→读计数器", "定时器继续计数(仿真环境可能无法完全模拟，但需记录)", "仿真环境受限"),

    # ===================================================
    # 6. 中断与复位请求
    # ===================================================
    ("中断与复位请求", "HPI路由", "HPI单核路由验证(C0/C1独立)", "ermu_hpi_test", "确认HPI可独立路由到Core0或Core1", "HPIE_C0=1,HPIE_C1=0→HPI响应错误→读hpi_irq_o; 再交换", "hpi_irq_o[0]=1,hpi_irq_o[1]=0; 反之亦然", "待实现"),
    ("中断与复位请求", "HPI路由", "HPI双核并发路由验证", "ermu_hpi_test", "确认HPI可同时路由到双核", "HPIE=11→HPI响应错误→读hpi_irq_o", "hpi_irq_o[1:0]=11", "待实现"),
    ("中断与复位请求", "HPI路由", "HPI作为NMI直接送达处理器", "ermu_hpi_test", "确认HPI不经过INTC，直接作为NMI", "触发HPI→监测hpi_irq_o时序", "hpi_irq_o无需INTC使能即产生", "待实现"),

    ("中断与复位请求", "LPI中断", "LPI三路独立中断验证", "ermu_error_rsp_test", "确认LPI0/1/2三路独立产生", "分别配置ERC=LPI0/LPI1/LPI2→注入错误→监测lpi_irq_o", "lpi_irq_o[0/1/2]各自独立触发", "待实现"),
    ("中断与复位请求", "LPI中断", "LPI需INTC使能才有效", "ermu_error_rsp_test", "确认LPI需要INTC配合才能使系统级中断生效", "配LPI→注入错误→监测lpi_irq_o", "lpi_irq_o=1(ERMU级可观测)", "待实现"),

    ("中断与复位请求", "复位请求", "应用复位请求(arst_req_o)产生", "ermu_error_rsp_test", "确认APP_RST错误响应产生应用复位请求", "配ERC=0x6→注入错误→监测arst_req_o", "arst_req_o=1", "待实现"),
    ("中断与复位请求", "复位请求", "系统复位请求(srst_req_o)产生", "ermu_error_rsp_test", "确认SYS_RST错误响应产生系统复位请求", "配ERC=0x7→注入错误→监测srst_req_o", "srst_req_o=1", "待实现"),
    ("中断与复位请求", "复位请求", "复位请求脉冲宽度验证", "ermu_error_rsp_test", "确认复位请求信号保持足够宽度", "触发复位请求→测量srst_req_o/arst_req_o脉宽", "复位请求信号宽度≥1个pclk周期", "待实现"),

    # ===================================================
    # 7. 紧急停止(EMS)
    # ===================================================
    ("紧急停止(EMS)", "同步模式(EMSMOD=0)", "上升沿触发→EMSS=1→软件清除", "ermu_pssr_test", "确认同步模式上升沿触发后EMSS=1，可通过EMSCLR清除", "配EMSEN=1,EMSMOD=0,EMSPOL=0→给EMSP_IN上升沿→读EMSS→写EMSCLR→读EMSS", "上升沿后EMSS=1; EMSCLR后EMSS=0", "待实现(DUT IMP_EMS=1支持)"),
    ("紧急停止(EMS)", "同步模式(EMSMOD=0)", "下降沿触发→EMSS=1→软件清除", "ermu_pssr_test", "确认同步模式下降沿触发", "配EMSEN=1,EMSMOD=0,EMSPOL=1→给EMSP_IN下降沿→读EMSS", "下降沿后EMSS=1", "待实现"),
    ("紧急停止(EMS)", "同步模式(EMSMOD=0)", "EMSSET软件触发", "ermu_pssr_test", "确认写EMSSET=1可在无端口过渡时软件触发EMS", "不操作EMSP_IN，直接写EMSSET=1→读EMSS", "EMSS=1(软件触发成功)", "待实现"),
    ("紧急停止(EMS)", "异步模式(EMSMOD=1)", "高电平激活、低电平释放", "ermu_pssr_test", "确认异步模式电平触发，不可软件清除", "配EMSEN=1,EMSMOD=1,EMSPOL=0→拉高EMSP_IN→读EMSS→写EMSCLR→读EMSS→拉低EMSP_IN→读EMSS", "高电平:EMSS=1,EMSCLR无效; 低电平:EMSS自动=0", "待实现"),
    ("紧急停止(EMS)", "异步模式(EMSMOD=1)", "低电平激活、高电平释放", "ermu_pssr_test", "确认异步模式极性反转", "配EMSEN=1,EMSMOD=1,EMSPOL=1→拉低EMSP_IN→读EMSS→拉高EMSP_IN→读EMSS", "低电平:EMSS=1; 高电平:EMSS自动=0", "待实现"),
    ("紧急停止(EMS)", "异步模式(EMSMOD=1)", "EMSSET在异步模式无效", "ermu_pssr_test", "确认异步模式软件触发无效", "EMSEN=1,EMSMOD=1→写EMSSET=1→读EMSS", "EMSS保持=0(EMSSET在异步模式被忽略)", "待实现"),
    ("紧急停止(EMS)", "端口选择", "EMSEL切换EMSP端口", "ermu_pssr_test", "确认可选择4个EMS端口之一", "分别配置EMSEL=00/01/10/11→在对应EMSP_IN操作→读EMSS", "只有被选中的EMSP_IN端口过渡触发EMS", "待实现"),
    ("紧急停止(EMS)", "极性控制", "EMSPOL控制输入极性", "ermu_pssr_test", "确认EMSPOL可翻转输入端口极性", "EMSPOL=0时上升沿触发; EMSPOL=1时下降沿触发", "极性选择正确", "待实现"),
    ("紧急停止(EMS)", "使能控制", "EMSEN=0时EMS功能禁用", "ermu_pssr_test", "确认EMSEN=0时所有EMS功能关闭", "EMSEN=0→操作EMSP_IN→读EMSS", "EMSS保持=0", "待实现"),
    ("紧急停止(EMS)", "EMS_EVENT输出", "EMS触发时ems_event输出验证", "ermu_pssr_test", "确认EMS触发时ems_event信号有效", "触发EMS→监测ems_event信号", "ems_event=1(对应错误ID=132)", "待实现"),

    # ===================================================
    # 8. PSSR端口安全状态请求
    # ===================================================
    ("PSSR", "触发源选择", "EOyS.EOS触发PSSR", "ermu_pssr_test", "确认EOS=1可触发PSSR(pssr_o=1)", "配PSSRS0=1(EOS0)→SET EOS→监测pssr_o", "pssr_o=1", "待实现"),
    ("PSSR", "触发源选择", "LPI触发PSSR", "ermu_pssr_test", "确认LPI可触发PSSR", "配PSSRS4(LPI0)=1→触发LPI→监测pssr_o", "pssr_o=1", "待实现"),
    ("PSSR", "触发源选择", "HPI触发PSSR", "ermu_pssr_test", "确认HPI可触发PSSR", "配PSSRS7(HPI)=1→触发HPI→监测pssr_o", "pssr_o=1", "待实现"),
    ("PSSR", "触发源选择", "复位请求触发PSSR", "ermu_pssr_test", "确认复位请求可触发PSSR", "配PSSRS8/9=1→触发APP_RST/SYS_RST→监测pssr_o", "pssr_o=1", "待实现"),
    ("PSSR", "触发源选择", "多触发源并发PSSR", "ermu_pssr_test", "确认多个PSSRS源同时使能时PSSR正常产生", "配多PSSRS=1→同时触发多个源→监测pssr_o", "pssr_o=1(多源OR逻辑)", "待实现"),
    ("PSSR", "禁用验证", "PSSRS=全0时PSSR不产生", "ermu_pssr_test", "确认PSSRS全0时任何条件不触发PSSR", "PSSRS=全0→触发各种源→监测pssr_o", "pssr_o保持=0", "待实现"),

    # ===================================================
    # 9. 复位行为
    # ===================================================
    ("复位行为", "上电复位(pvs)", "上电复位后全寄存器复位值验证", "ermu_reset_test", "确认rst_pvs_s复位后所有寄存器回到复位值", "触发rst_pvs_s→释放→读所有寄存器", "所有寄存器=复位值；CFGLOCK=1, PERLOCK=0, EOS=1", "待实现"),
    ("复位行为", "系统复位(pss_h2)", "系统复位后寄存器状态验证", "ermu_reset_test", "确认rst_pss_h2_s复位后状态", "操作中触发rst_pss_h2_s→释放→读关键寄存器", "大部分寄存器=复位值, ESS可能不确定(取决于电压检测)", "待实现"),
    ("复位行为", "应用复位(h2)", "应用复位后状态保持验证", "ermu_reset_test", "确认rst_h2_s复位类型的行为", "配置部分寄存器→触发rst_h2_s→释放→读寄存器", "根据复位行为表验证各寄存器受影响程度", "待实现"),
    ("复位行为", "操作中复位恢复", "定时器运行中复位恢复", "ermu_reset_test", "确认定时器计数期间复位能正确恢复", "启动定时器→触发复位→释放→读相关寄存器", "定时器停止，相关寄存器=复位值", "待实现"),
    ("复位行为", "操作中复位恢复", "错误处理中复位恢复", "ermu_reset_test", "确认错误处理过程中复位能正确恢复", "注入错误→在响应进行中触发复位→释放→读状态", "错误状态清除，系统恢复到复位状态", "待实现"),

    # ===================================================
    # 10. 压力与并发
    # ===================================================
    ("压力与并发", "多源并发", "10+错误源同时触发", "ermu_concurrent_test", "确认大量错误源同时触发时ESS全部正确记录", "同时注入10+个分布式错误源→读全部ESSj", "所有注入源ESS位=1", "待实现"),
    ("压力与并发", "多源并发", "不同响应类型错误源并发", "ermu_concurrent_test", "确认不同响应类型错误并发时各自正确响应", "同时注入NA/LPI/HPI/APP_RST/SYS_RST各类错误→监测各输出", "各类型响应信号均正确产生", "待实现"),

    ("压力与并发", "多通道并发", "4通道+4定时器并发", "ermu_concurrent_test", "确认4通道各自定时器同时运行时互不干扰", "4通道各配不同定时器参数→同时启动→读各通道状态", "各通道独立运行，输出正确", "待实现"),

    ("压力与并发", "随机化", "随机错误源ID+随机ERC编码+随机定时器参数", "ermu_stress_test", "确认随机化激励下系统稳定运行", "大量随机化测试(50+seed)→自动检查ESS/输出/中断", "无死锁、无意外复位、无状态错乱", "待实现"),
    ("压力与并发", "随机化", "288源全遍历随机覆盖", "ermu_stress_test", "确认所有288源在随机测试中被覆盖", "随机seed组合确保288源均被测试→检查覆盖率", "每个错误源至少被触发N次", "待实现"),

    ("压力与并发", "全功能", "全功能序列化综合测试", "ermu_func_full_test", "确认完整功能流程串联正确", "按应用手册流程: 读ESS→清EOS→配端口→配OM→配ERC→配定时器→设只读→处理中断/复位", "完整流程无错误，所有功能串联正确", "待实现"),
    ("压力与并发", "全功能", "长时间运行稳定性", "ermu_stress_test", "确认长时间运行无异常", "连续运行>100k个pclk周期，持续注入错误+清除→监控状态", "无内存泄漏、无状态累积错误", "待实现"),

    # ===================================================
    # 11. 特殊场景
    # ===================================================
    ("特殊场景", "双时钟域", "pclk与clk_hrc异步跨域验证", "ermu_concurrent_test", "确认APB访问(pclk域)与定时器计数(clk_hrc域)正确跨域", "高频APB访问+定时器运行→检查数据一致性", "APB读写正确，定时器计数正确，无亚稳态", "待实现"),
    ("特殊场景", "双时钟域", "APB读写与定时器计数竞争", "ermu_concurrent_test", "确认APB访问定时器寄存器时不会读取到错误中间值", "定时器计数中→高频读CTCNT/TTCNT/WTCNT→验证值单调递增", "CNT值始终≥上一次读取值，无回退", "待实现"),

    ("特殊场景", "DFT模式", "DFT=1时功能不可见", "ermu_reg_test", "确认DFT信号=1时功能模式操作被隔离", "scan_mode=1→尝试寄存器访问→验证行为", "DFT模式下功能寄存器行为符合DFT规格(如读回固定值)", "功能验证仅确认DFT=0正常工作"),

    ("特殊场景", "APB错误响应", "未实现地址访问返回pslverr", "ermu_reg_test", "确认访问保留/未实现地址时APB返回pslverr", "读/写超出寄存器映射范围的地址→监测pslverr", "pslverr=1", "待实现"),
    ("特殊场景", "APB错误响应", "非字写返回pslverr", "ermu_reg_test", "确认非字写访问时APB返回错误", "尝试pstrb≠4'b1111的写操作→监测pslverr", "pslverr=1或写操作被忽略", "待实现"),
]

# ── Write rows ──────────────────────────────────────
row_num = 2
cat_start = {}   # track start row per category for merging
prev_cat = None
seq = 1

for i, (cat, sub, item, case, goal, method, expect, note) in enumerate(rows):
    r = row_num + i
    if cat != prev_cat:
        if prev_cat and cat_start.get(prev_cat):
            sc = cat_start[prev_cat]
            if r - 1 > sc:
                ws.merge_cells(start_row=sc, start_column=2, end_row=r-1, end_column=2)
        cat_start[cat] = r
        prev_cat = cat

    ws.cell(row=r, column=1, value=seq).font = cell_font
    ws.cell(row=r, column=1).alignment = center_align
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=2, value=cat).font = cell_font
    ws.cell(row=r, column=2).alignment = cell_align
    ws.cell(row=r, column=2).border = thin_border
    ws.cell(row=r, column=3, value=sub).font = cell_font
    ws.cell(row=r, column=3).alignment = cell_align
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=4, value=item).font = cell_font
    ws.cell(row=r, column=4).alignment = cell_align
    ws.cell(row=r, column=4).border = thin_border
    ws.cell(row=r, column=5, value=case).font = cell_font
    ws.cell(row=r, column=5).alignment = cell_align
    ws.cell(row=r, column=5).border = thin_border
    ws.cell(row=r, column=6, value=goal).font = cell_font
    ws.cell(row=r, column=6).alignment = cell_align
    ws.cell(row=r, column=6).border = thin_border
    ws.cell(row=r, column=7, value=method).font = cell_font
    ws.cell(row=r, column=7).alignment = cell_align
    ws.cell(row=r, column=7).border = thin_border
    ws.cell(row=r, column=8, value=expect).font = cell_font
    ws.cell(row=r, column=8).alignment = cell_align
    ws.cell(row=r, column=8).border = thin_border
    ws.cell(row=r, column=9, value="待验证").font = cell_font
    ws.cell(row=r, column=9).alignment = center_align
    ws.cell(row=r, column=9).border = thin_border
    ws.cell(row=r, column=10, value=note).font = cell_font
    ws.cell(row=r, column=10).alignment = cell_align
    ws.cell(row=r, column=10).border = thin_border

    seq += 1

# merge last category
last_r = row_num + len(rows) - 1
if prev_cat and cat_start.get(prev_cat):
    sc = cat_start[prev_cat]
    if last_r > sc:
        ws.merge_cells(start_row=sc, start_column=2, end_row=last_r, end_column=2)

# Apply light-blue fill to category column
for rr in range(2, row_num + len(rows)):
    cell = ws.cell(row=rr, column=2)
    cell.fill = cat_fill

# alternate row shading for readability
light_gray = PatternFill("solid", fgColor="F2F2F2")
for rr in range(2, row_num + len(rows)):
    if (rr % 2) == 0:
        for cc in range(1, 11):
            c = ws.cell(row=rr, column=cc)
            if cc != 2:  # don't override category fill
                c.fill = light_gray

# row height
for rr in range(2, row_num + len(rows)):
    ws.row_dimensions[rr].height = 56

ws.row_dimensions[1].height = 28

output = "d:/mywork/vsc000/ERMU_UT/doc/ermu_verification_plan.xlsx"
wb.save(output)
print(f"Saved {len(rows)} verification items to {output}")
