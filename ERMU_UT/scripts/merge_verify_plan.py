import json, copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import OrderedDict

# ── Load data ──────────────────────────────────────
with open(r"d:\mywork\vsc000\ERMU_UT\doc\corrections.json", "r", encoding="utf-8") as f:
    corrections = json.load(f)
with open(r"d:\mywork\vsc000\ERMU_UT\doc\additions.json", "r", encoding="utf-8") as f:
    additions = json.load(f)

# ── Read original Excel ─────────────────────────────
wb = openpyxl.load_workbook(r"d:\mywork\vsc000\ERMU_UT\doc\ermu_verification_plan.xlsx")
ws = wb["验证计划"]

# Build row map: 序号 -> {col_name: cell_value, row_num: excel_row}
col_map = {}
for c in range(1, 11):
    col_map[ws.cell(row=1, column=c).value] = c
rev_col_map = {v: k for k, v in col_map.items()}

orig_data = []
for r in range(2, ws.max_row + 1):
    seq = ws.cell(row=r, column=col_map["序号"]).value
    if seq is None:
        continue
    row_dict = {}
    for col_name, col_idx in col_map.items():
        row_dict[col_name] = ws.cell(row=r, column=col_idx).value
    row_dict["_excel_row"] = r
    orig_data.append(row_dict)

print(f"Original: {len(orig_data)} rows")

# ── Apply corrections ───────────────────────────────
corr_count = 0
for c in corrections:
    target_seq = c.get("序号")
    col_name = c.get("col")
    new_val = c.get("new")
    for row in orig_data:
        if row["序号"] == target_seq:
            if row.get(col_name) != new_val:
                row[col_name] = new_val
                corr_count += 1
                print(f"  Corrected 序号{target_seq} [{col_name}]: {str(row.get(col_name))[:60]} -> {str(new_val)[:60]}")
            break
print(f"Applied {corr_count} corrections")

# ── Build category order map ────────────────────────
cat_order = ["寄存器访问验证", "错误源管理", "错误响应控制", "错误输出通道",
             "定时器功能", "中断与复位请求", "紧急停止(EMS)", "PSSR",
             "复位行为", "压力与并发", "特殊场景"]

# ── Insert additions into correct categories ────────
# Group additions by category, find insertion point after last item in same category
cat_last_idx = {}
for i, row in enumerate(orig_data):
    cat_last_idx[row["功能大项"]] = i

additions_by_cat = OrderedDict()
for a in additions:
    cat = a["功能大项"]
    if cat not in additions_by_cat:
        additions_by_cat[cat] = []
    additions_by_cat[cat].append(a)

# Insert in reverse order (so indices don't shift)
inserted = 0
for cat in reversed(list(additions_by_cat.keys())):
    items = additions_by_cat[cat]
    insert_pos = cat_last_idx.get(cat, len(orig_data) - 1) + 1
    for item in reversed(items):
        orig_data.insert(insert_pos, item)
        inserted += 1
    print(f"  Added {len(items)} items to '{cat}' at position {insert_pos}")
print(f"Total inserted: {inserted} items")

# ── Renumber ────────────────────────────────────────
for i, row in enumerate(orig_data):
    row["序号"] = i + 1

print(f"Final: {len(orig_data)} rows")

# ── Rebuild Excel ───────────────────────────────────
# Styles
hdr_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")
hdr_fill   = PatternFill("solid", fgColor="4472C4")
hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font  = Font(name="Arial", size=10)
cell_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))
cat_fill   = PatternFill("solid", fgColor="D6E4F0")
new_fill   = PatternFill("solid", fgColor="E2EFDA")  # light green for new items
corr_fill  = PatternFill("solid", fgColor="FCE4D6")  # light orange for corrected items
light_gray = PatternFill("solid", fgColor="F2F2F2")

col_widths = {"A": 6, "B": 18, "C": 26, "D": 32, "E": 26, "F": 40, "G": 62, "H": 40, "I": 14, "J": 22}

# Create new workbook
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "验证计划"

for col_letter, w in col_widths.items():
    ws2.column_dimensions[col_letter].width = w

# Header
headers = ["序号", "功能大项", "功能子项", "验证项目", "验证Case名称", "验证目标", "验证方法与条件", "期待结果", "验证结果", "备注"]
for c, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = thin_border

ws2.freeze_panes = "A2"

# Track corrected seq numbers and new item seq numbers for highlighting
corrected_seqs = set()
for c in corrections:
    corrected_seqs.add(c.get("序号"))
new_items_count = 0

# Write data
prev_cat = None
cat_start_row = 2
for i, row in enumerate(orig_data):
    r = i + 2  # Excel row (1-indexed, row 1 is header)
    seq = row["序号"]
    cat = row["功能大项"]

    is_new = "审查新增" in str(row.get("备注", ""))
    is_corrected = seq in corrected_seqs and not is_new
    if is_new:
        new_items_count += 1

    # Category merge tracking
    if cat != prev_cat:
        if prev_cat and cat_start_row < r - 1:
            ws2.merge_cells(start_row=cat_start_row, start_column=2, end_row=r-1, end_column=2)
        cat_start_row = r
        prev_cat = cat

    # Determine row fill
    if is_new:
        row_fill = new_fill
    elif is_corrected:
        row_fill = corr_fill
    elif i % 2 == 0:
        row_fill = light_gray
    else:
        row_fill = None

    for ci, col_name in enumerate(headers, 1):
        val = row.get(col_name, "待验证" if col_name == "验证结果" else "")
        cell = ws2.cell(row=r, column=ci, value=val)
        cell.font = cell_font
        cell.alignment = center_align if col_name in ("序号", "验证结果") else cell_align
        cell.border = thin_border
        if row_fill and col_name != "功能大项":
            cell.fill = row_fill
        elif col_name == "功能大项":
            cell.fill = cat_fill

    ws2.row_dimensions[r].height = 60

# Final category merge
last_r = 2 + len(orig_data) - 1
if prev_cat and cat_start_row < last_r:
    ws2.merge_cells(start_row=cat_start_row, start_column=2, end_row=last_r, end_column=2)

ws2.row_dimensions[1].height = 28

# ── Add legend sheet ────────────────────────────────
ws3 = wb2.create_sheet("图例")
ws3['A1'] = "颜色图例"
ws3['A1'].font = Font(name="Arial", bold=True, size=12)
ws3['A3'] = "浅橙色行"
ws3['A3'].fill = corr_fill
ws3['B3'] = "已修正项 (基于审查结果修改了验证方法/期待结果)"
ws3['A4'] = "浅绿色行"
ws3['A4'].fill = new_fill
ws3['B4'] = "新增项 (基于审查结果新补充的验证项)"
ws3['A5'] = "浅蓝色列"
ws3['A5'].fill = cat_fill
ws3['B5'] = "功能大项分组标识"
ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 60

# ── Add summary sheet ───────────────────────────────
ws4 = wb2.create_sheet("统计")
ws4['A1'] = "验证计划统计"
ws4['A1'].font = Font(name="Arial", bold=True, size=14)

cat_counts = OrderedDict()
for row in orig_data:
    cat = row["功能大项"]
    cat_counts[cat] = cat_counts.get(cat, 0) + 1

ws4['A3'] = "功能大项"; ws4['B3'] = "验证项数量"
ws4['A3'].font = Font(name="Arial", bold=True); ws4['B3'].font = Font(name="Arial", bold=True)
r = 4
for cat in cat_order:
    if cat in cat_counts:
        ws4.cell(row=r, column=1, value=cat)
        ws4.cell(row=r, column=2, value=cat_counts[cat])
        r += 1
ws4.cell(row=r, column=1, value="合计").font = Font(name="Arial", bold=True)
ws4.cell(row=r, column=2, value=len(orig_data)).font = Font(name="Arial", bold=True)
ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 14

output = r"d:\mywork\vsc000\ERMU_UT\doc\ermu_verification_plan.xlsx"
wb2.save(output)
print(f"\nSaved: {output}")
print(f"Original: {len(orig_data) - inserted} | Corrected: {len(corrected_seqs)} | New: {new_items_count} | Total: {len(orig_data)}")
