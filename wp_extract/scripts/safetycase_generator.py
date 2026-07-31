"""
Safety Case Generator Module
Reads a Safety Case template, searches for work products in a directory,
and fills in the template to produce a project-specific safety case report.

Template: Safety case_template.xlsx (read-only, never modified).
Output: <project>_safetycase.xlsx (copy of template with cells filled).
"""

import os
import re
import json
import sys
from datetime import datetime
from copy import copy
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.worksheet.hyperlink import Hyperlink

from file_matcher import MatchResult, match_keyword_to_file, discover_files, load_synonym_config
from metadata_extractor import extract_metadata, DocumentMetadata, NOT_FOUND

# Light red fill for unmatched cells
UNMATCHED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# D column entries that should only get directory paths (not individual filenames)
DIRECTORY_ONLY_ENTRIES = {
    '5.5.1　Organization-specific rules and processes for functional safety',
    '6.5.2　Hardware-software interface specification (HSI)',
    '7.5.1   Hardware design specification',
    '7.5.3   Hardware design verification report',
    '10.5.1 Hardware integration and verification specification',
    '10.5.2 Hardware integration and verification report resulting from requirements',
    '9.5.2  Verification specification',
    '10.5.2 Documentation guideline requirements',
    '13.5.1 Hardware element evaluation plan',
}

# D column entries that should NOT be filled (skip H-L entirely)
SKIP_ENTRIES = {
    '5.5.4 Identified safety anomaly reports',
    '5.5.5 Supply agreement',
    '8.5.2 Change request',
    '8.5.3 Impact analysis and change request plan',
    '8.5.4 Change report',
}

# Special row rules: D column pattern → (directory, team, ref_pattern)
# ref_pattern is used to find a reference file for metadata extraction
SPECIAL_RULES = [
    {
        'd_pattern': '',
        'e_pattern': 'User Manuals|Target specification',
        'directory': 'part5/手册/',
        'team': 'DesignTeam',
        'ref_pattern': r'总线|bus|BUS',
        'force_status': 'Released',
    },
    {
        'd_pattern': '7.5.2 Hardware safety analysis report',
        'e_pattern': 'DFMEA',
        'keywords': ['DFMEA'],
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '9.5.2 Verification specification',
        'e_pattern': '',
        'directory': 'part5/PAC_03_FS-2_HW_verification_plan/',
        'team': 'VerificationTeam',
        'ref_pattern': r'总线|bus|BUS',
    },
    {
        'd_pattern': '9.5.3 Verification report',
        'e_pattern': '',
        'directory': 'part5/PAC_03_FS-7_HW_verification_pre_tape_out/',
        'team': 'VerificationTeam',
        'ref_pattern': r'总线|bus|BUS',
    },
    {
        'd_pattern': '6.5.1 Update of the ASIL attribute',
        'e_pattern': '',
        'keywords': ['SEooC definition', 'SEooC'],
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '7.5.1 Dependent Failures Analysis',
        'e_pattern': '',
        'keywords': ['DFA'],
        'prefer_pattern': r'C\d+_DFA\.(xlsx|xlsm)',
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '7.5.2 Dependent Failures Analysis Verification Report',
        'e_pattern': '',
        'keywords': ['Qualitative safety analysis and DFA verification review'],
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '8.5.1 Safety analyses',
        'e_pattern': '',
        'keywords': ['FMEDA', 'Quantitative safety analysis'],
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '8.5.2 Safety analyses verification report',
        'e_pattern': '',
        'keywords': ['Quantitative safety analysis verification review'],
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '5.5.2 Evidence of competence management',
        'e_pattern': '',
        'keywords': ['Safety plan'],
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '6.5.3 Hardware safety requirements verification report',
        'e_pattern': '',
        'keywords': ['HSR specification verification review'],
        'team': 'DesignTeam',
    },
    {
        'd_pattern': '9 Safety Element out of Context',
        'e_pattern': 'Functional safety manual',
        'keywords': ['safety manual'],
        'team': 'FusaTeam',
        'force_status': 'Released',
    },
]

CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _norm_d_column(value) -> str:
    """Normalize D column text for comparison (collapse whitespace)."""
    if not value:
        return ''
    return ' '.join(str(value).split())


def _find_ref_file(directory: str, pattern: str, all_files: list[str]) -> Optional[str]:
    """Find a file in directory matching the pattern. Used for reference metadata lookup."""
    import re as _re
    pat = _re.compile(pattern)
    norm_dir = directory.replace('\\', '/').lower()
    candidates = []
    for fp in all_files:
        fn = os.path.basename(fp)
        fp_norm = fp.replace('\\', '/').lower()
        if norm_dir in fp_norm and pat.search(fn):
            candidates.append(fp)
    return candidates[0] if candidates else None


def _match_special_rule(d_value: str, e_value: str) -> Optional[dict]:
    """Check if this row matches a special mapping rule."""
    d_norm = _norm_d_column(d_value)
    e_norm = _norm_d_column(e_value)
    for rule in SPECIAL_RULES:
        if rule['d_pattern']:
            if _norm_d_column(rule['d_pattern']) not in d_norm:
                continue
        if rule['e_pattern']:
            parts = [_norm_d_column(p) for p in rule['e_pattern'].split('|')]
            if not any(p in e_norm for p in parts):
                continue
        return rule
    return None


def _is_skip_entry(d_value) -> bool:
    """Check if this D column entry should be skipped entirely."""
    norm = _norm_d_column(d_value)
    for entry in SKIP_ENTRIES:
        if _norm_d_column(entry) in norm:
            return True
    return False


def _is_directory_only_entry(d_value) -> bool:
    """Check if this D column entry should only get directory paths."""
    norm = _norm_d_column(d_value)
    for entry in DIRECTORY_ONLY_ENTRIES:
        if _norm_d_column(entry) in norm:
            return True
    return False


def _get_directory_path(filepath: str, doc_dir: str) -> str:
    """Extract the directory path (not the filename) from a file path."""
    if doc_dir and filepath.startswith(doc_dir):
        rel = os.path.relpath(filepath, doc_dir)
        dir_part = os.path.dirname(rel)
        return dir_part if dir_part else rel
    return os.path.dirname(filepath)


def _set_h_cell(ws, row: int, display_text: str, doc_dir: str,
                file_paths: list[str] = None, dir_paths: list[str] = None):
    """
    Write display_text to H column and attach a hyperlink.
    - Single or multiple files: link to the first file's absolute path.
    - Directory-only entries: link to the first directory's absolute path.
    - display_text may contain multiple lines (\\n-separated).
    """
    import openpyxl.worksheet.hyperlink as _hl

    cell = ws.cell(row=row, column=8)
    cell.value = display_text

    target = None
    if file_paths:
        target = file_paths[0]
    elif dir_paths:
        target = os.path.join(doc_dir, dir_paths[0]) if doc_dir else dir_paths[0]

    if target and os.path.exists(target):
        normalized = target.replace('\\', '/')
        cell.hyperlink = _hl.Hyperlink(ref=cell.coordinate, target=normalized)


def _load_project_meta(config_path: str = None) -> dict:
    """Load project_meta from synonym config."""
    path = config_path or os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                        '..', 'settings', 'synonym_config.json')
    try:
        with open(path, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        return cfg.get('project_meta', {})
    except Exception:
        return {}


def print_directory_tree(directory: str, max_depth: int = 5) -> str:
    """
    Build a tree-like string representation of a directory.
    Skips hidden dirs, temp files, and __pycache__.
    """
    lines = [os.path.basename(directory) + '/']

    def _walk(dirpath: str, prefix: str, depth: int):
        if depth > max_depth:
            lines.append(prefix + '...')
            return
        try:
            entries = sorted(os.listdir(dirpath))
        except PermissionError:
            return

        # Filter: skip hidden, temp, pycache
        entries = [e for e in entries
                   if not e.startswith('.')
                   and not e.startswith('~$')
                   and e != '__pycache__']

        dirs = [e for e in entries if os.path.isdir(os.path.join(dirpath, e))]
        files = [e for e in entries if os.path.isfile(os.path.join(dirpath, e))]

        for i, d in enumerate(dirs):
            is_last = (i == len(dirs) - 1) and not files
            connector = '└── ' if is_last else '├── '
            lines.append(prefix + connector + d + '/')
            next_prefix = prefix + ('    ' if is_last else '│   ')
            _walk(os.path.join(dirpath, d), next_prefix, depth + 1)

        for i, f in enumerate(files):
            is_last = (i == len(files) - 1)
            connector = '└── ' if is_last else '├── '
            lines.append(prefix + connector + f)

    _walk(directory, '', 1)
    return '\n'.join(lines)


def _restore_drawings(template_path: str, output_path: str):
    """
    openpyxl drops drawings AND their references (Content_Types, sheet rels).
    Restore everything drawing-related from template.
    """
    import zipfile as _zipfile
    import shutil as _shutil
    import re as _re

    # Identify what was lost
    tpl_names = set()
    with _zipfile.ZipFile(template_path, 'r') as z:
        tpl_names = set(z.namelist())
    out_names = set()
    with _zipfile.ZipFile(output_path, 'r') as z:
        out_names = set(z.namelist())

    # 1) Missing drawing/vml files
    missing_files = {}
    for name in tpl_names:
        if ('drawings/' in name or 'vmlDrawing' in name) and name not in out_names:
            missing_files[name] = True

    # 2) Missing sheet rels that reference drawings
    missing_rels = {}
    for name in tpl_names:
        if 'worksheets/_rels/sheet' in name and name not in out_names:
            missing_rels[name] = True

    # 3) Content_Types.xml — openpyxl strips drawing Override entries
    ct_missing_entries = []
    with _zipfile.ZipFile(template_path, 'r') as z:
        tpl_ct = z.read('[Content_Types].xml').decode('utf-8')
        # Simple pattern: extract all PartName values
        tpl_all = _re.findall(r'PartName=\"([^\"]*)\"', tpl_ct)
        tpl_drawing_refs = [r for r in tpl_all if 'drawing' in r.lower()]
    with _zipfile.ZipFile(output_path, 'r') as z:
        out_ct = z.read('[Content_Types].xml').decode('utf-8')
    for ref in tpl_drawing_refs:
        if ref not in out_ct:
            ct_missing_entries.append(
                f'<Override PartName="{ref}" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
            )

    if not missing_files and not missing_rels and not ct_missing_entries:
        return

    # Rebuild output ZIP with all missing pieces
    tmp_path = output_path + '.tmp'
    with _zipfile.ZipFile(template_path, 'r') as tpl_z:
        with _zipfile.ZipFile(output_path, 'r') as out_z:
            with _zipfile.ZipFile(tmp_path, 'w', _zipfile.ZIP_DEFLATED) as tmp_z:
                for item in out_z.infolist():
                    data = out_z.read(item.filename)
                    name = item.filename
                    # Patch Content_Types.xml to add drawing refs
                    if name == '[Content_Types].xml' and ct_missing_entries:
                        text = data.decode('utf-8')
                        insert_pos = text.rfind('</Types>')
                        if insert_pos > 0:
                            text = text[:insert_pos] + ''.join(ct_missing_entries) + text[insert_pos:]
                        data = text.encode('utf-8')
                    tmp_z.writestr(item, data)
                # Inject missing files from template
                for name in missing_files:
                    tmp_z.writestr(name, tpl_z.read(name))
                for name in missing_rels:
                    tmp_z.writestr(name, tpl_z.read(name))

    _shutil.move(tmp_path, output_path)


def _read_old_safety_case(output_path: str) -> dict:
    """Read old Safety Case sheet into {(row, col): value} dict. Returns {} if not found."""
    import openpyxl as _xl
    try:
        wb = _xl.load_workbook(output_path, data_only=True)
    except Exception:
        return {}
    if 'Safety Case' not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb['Safety Case']
    old_data = {}
    for r in range(14, ws.max_row + 1):
        for c in [8, 9, 10, 11, 12]:
            old_data[(r, c)] = str(ws.cell(row=r, column=c).value or '').strip()
    wb.close()
    return old_data


def _read_old_revision_full(output_path: str) -> tuple[Optional[str], list[dict]]:
    """Read all revision rows from old report. Returns (latest_rev, [row_data, ...])."""
    import openpyxl as _xl
    import re as _re
    rows = []
    latest = None
    try:
        wb = _xl.load_workbook(output_path, data_only=True)
        if 'Revision History' in wb.sheetnames:
            ws = wb['Revision History']
            for r in range(3, ws.max_row + 1):
                b_val = ws.cell(row=r, column=2).value
                if b_val is not None and _re.match(r'^\d+\.\d+$', str(b_val).strip()):
                    latest = str(b_val).strip()
                    rows.append({
                        'rev': latest,
                        'date': str(ws.cell(row=r, column=3).value or '').strip(),
                        'author': str(ws.cell(row=r, column=4).value or '').strip(),
                        'desc': str(ws.cell(row=r, column=5).value or '').strip(),
                    })
        wb.close()
    except Exception:
        pass
    return latest, rows


def _append_revision_history(ws_rev, old_rev_rows: list[dict], new_rev: str,
                             fsm_name: str):
    """
    Fill Revision History sheet.
    - Copy old version rows first (incremental mode).
    - Then append new_rev row.
    """
    new_date = datetime.now().strftime('%Y-%m-%d')
    new_desc = 'Updated with latest info of Workproducts.'

    # Write old rows first
    for i, row_data in enumerate(old_rev_rows):
        r = 3 + i
        ws_rev.cell(row=r, column=2, value=row_data['rev'])
        ws_rev.cell(row=r, column=3, value=row_data['date'])
        ws_rev.cell(row=r, column=4, value=row_data['author'])
        ws_rev.cell(row=r, column=5, value=row_data['desc'])

    # Append new revision row
    new_row = 3 + len(old_rev_rows)
    ws_rev.cell(row=new_row, column=2, value=new_rev)
    ws_rev.cell(row=new_row, column=3, value=new_date)
    ws_rev.cell(row=new_row, column=4, value=fsm_name)
    ws_rev.cell(row=new_row, column=5, value=new_desc)

    # If Excel Table, expand range to cover the new row
    if ws_rev.tables:
        table = list(ws_rev.tables.values())[0]
        old_ref = table.ref
        parts = old_ref.split(':')
        if len(parts) == 2:
            col_letter = ''.join(c for c in parts[1] if c.isalpha())
            old_end = int(''.join(c for c in parts[1] if c.isdigit()))
            new_end = max(old_end, new_row)
            table.ref = f'{parts[0]}:{col_letter}{new_end}'


def _apply_incremental_diff(ws_new, old_data: dict):
    """
    Compare new Safety Case values against old data.
    - New/changed cells → light green
    - NA cells → light red (takes priority)
    """
    GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    for r in range(14, ws_new.max_row + 1):
        for c in [8, 9, 10, 11, 12]:
            new_cell = ws_new.cell(row=r, column=c)
            new_val = str(new_cell.value or '').strip()
            old_val = old_data.get((r, c), '')

            # NA always wins (red)
            if new_val == 'NA':
                new_cell.fill = UNMATCHED_FILL
                continue

            # New or changed → green
            if not old_val or old_val == 'NA' or new_val != old_val:
                if new_val:
                    new_cell.fill = GREEN_FILL


def fill_safety_case_report(
    template_path: str,
    doc_dir: str,
    output_path: str,
    config_path: str = None,
    min_score: float = 0.5,
    fresh: bool = True,
) -> str:
    """
    Read the Safety Case template, search for work products, fill in metadata,
    and save as a project-specific safety case report.

    Args:
        template_path: Path to Safety case_template.xlsx (read-only).
        doc_dir: Directory to search for work products.
        output_path: Output Excel path for the filled report.
        config_path: Path to synonym config (for project_meta).
        min_score: Keyword matching threshold.
        fresh: If True, copy from template (new). If False, update existing file (incremental).

    Returns:
        Path to the generated report.
    """
    # Load project metadata
    project_meta = _load_project_meta(config_path)
    project_name = project_meta.get('project', 'Unknown')
    fsm_name = project_meta.get('FSM', '')
    approver_name = project_meta.get('Approver', '')

    # Read old report BEFORE template copy (for incremental diff + version history)
    old_safety_case = {}
    old_latest_rev = None
    old_rev_rows = []
    if not fresh and os.path.exists(output_path):
        old_safety_case = _read_old_safety_case(output_path)
        old_latest_rev, old_rev_rows = _read_old_revision_full(output_path)

    # Always start from fresh template copy
    import shutil
    shutil.copy2(template_path, output_path)

    # Open the output for editing
    wb = openpyxl.load_workbook(output_path)

    # Sheets to NEVER modify: Safety Case Argument, Template Revision History
    # (we only touch Title, Revision History, Safety Case)

    # --- Discover files once ---
    all_files = discover_files(doc_dir)
    metadata_cache: dict[str, DocumentMetadata] = {}

    # --- Sheet: Title ---
    ws_title = wb['Title']
    # C5: replace <project>
    c5_val = str(ws_title['C5'].value or '')
    ws_title['C5'].value = c5_val.replace('<project>', project_name)

    # C6: will be filled after Revision History processing

    # C8: FSM
    if fsm_name:
        ws_title['C8'].value = fsm_name

    # C10: Approver
    if approver_name:
        ws_title['C10'].value = approver_name

    # C12: Date = today
    ws_title['C12'].value = datetime.now().strftime('%Y-%m-%d')

    # --- Sheet: Revision History ---
    ws_rev = wb['Revision History']
    import re as _re
    if old_latest_rev:
        latest_rev = old_latest_rev
    else:
        latest_rev = '0.00'
    try:
        rev_num = float(latest_rev)
        new_rev = f'{rev_num + 0.01:.2f}'
    except ValueError:
        new_rev = '0.01'

    _append_revision_history(ws_rev, old_rev_rows, new_rev, fsm_name)

    # Update Title C6 with the new revision
    ws_title['C6'].value = new_rev

    # --- Sheet: Safety Case ---
    ws_sc = wb['Safety Case']

    unmatched_count = 0
    matched_count = 0
    total_processed = 0

    for r in range(14, ws_sc.max_row + 1):
        f_val = str(ws_sc.cell(row=r, column=6).value or '').strip()  # F: Tailoring?
        if f_val.lower() != 'no':
            continue  # skip tailored-out rows

        total_processed += 1

        # Preserve existing H column values (reference data from template)
        existing_h = str(ws_sc.cell(row=r, column=8).value or '').strip()
        if existing_h and existing_h != '-':
            matched_count += 1
            continue  # keep template pre-filled row as-is

        e_val = str(ws_sc.cell(row=r, column=5).value or '').strip()  # E: keywords
        d_val = str(ws_sc.cell(row=r, column=4).value or '').strip()  # D: ISO chapter

        # Rule: skip certain D column entries entirely
        if _is_skip_entry(d_val):
            continue

        if not e_val:
            continue

        # Check for special mapping rules first
        special = _match_special_rule(d_val, e_val)
        if special:
            matched_count += 1
            ws_sc.cell(row=r, column=9).value = special.get('team', '')  # I: team

            # Type 1: keyword-based rules (search doc_dir for matching files)
            if special.get('keywords'):
                all_matched = []
                for kw in special['keywords']:
                    result = match_keyword_to_file(kw, all_files, min_score)
                    if result.is_matched:
                        for fp in result.matched_files:
                            if fp not in all_matched:
                                all_matched.append(fp)

                # Apply prefer_pattern filter if specified
                prefer_pattern = special.get('prefer_pattern')
                if prefer_pattern and all_matched:
                    import re as _re3
                    pp = _re3.compile(prefer_pattern)
                    filtered = [fp for fp in all_matched if pp.search(os.path.basename(fp))]
                    if filtered:
                        all_matched = filtered

                if all_matched:
                    first_file = all_matched[0]
                    if first_file not in metadata_cache:
                        metadata_cache[first_file] = extract_metadata(first_file)
                    meta = metadata_cache[first_file]
                    paths = []
                    for fp in all_matched:
                        if doc_dir and fp.startswith(doc_dir):
                            paths.append(os.path.relpath(fp, doc_dir))
                        else:
                            paths.append(fp)
                    _set_h_cell(ws_sc, r, '\n'.join(paths), doc_dir,
                                file_paths=all_matched)
                    ws_sc.cell(row=r, column=10).value = (
                        special['force_status'] if special.get('force_status')
                        else meta.status
                    )
                    ws_sc.cell(row=r, column=11).value = meta.revision      # K
                    ws_sc.cell(row=r, column=12).value = meta.date          # L
                    # Apply NA fill for any NA metadata fields
                    for col in [11, 12]:
                        cell = ws_sc.cell(row=r, column=col)
                        if cell.value == 'NA':
                            cell.fill = UNMATCHED_FILL
                else:
                    for col in [8, 11, 12]:
                        ws_sc.cell(row=r, column=col).value = 'NA'
                        ws_sc.cell(row=r, column=col).fill = UNMATCHED_FILL
                    if special.get('force_status'):
                        ws_sc.cell(row=r, column=10).value = special['force_status']
                continue

            # Type 2: directory-based rules (fill directory path + ref metadata)
            dir_path = special.get('directory', '')
            _set_h_cell(ws_sc, r, dir_path, doc_dir, dir_paths=[dir_path] if dir_path else [])

            ref_file = None
            if special.get('prefer_file'):
                for fp in all_files:
                    fn = os.path.basename(fp)
                    fp_norm = fp.replace('\\', '/').lower()
                    if special['prefer_file'].lower() in fn.lower() and dir_path.replace('\\', '/').lower() in fp_norm:
                        ref_file = fp
                        break
            if not ref_file and special.get('ref_pattern'):
                norm_dir = dir_path.replace('\\', '/').lower().rstrip('/')
                import re as _re2
                pat = _re2.compile(special['ref_pattern'])
                for fp in all_files:
                    fp_norm = fp.replace('\\', '/').lower()
                    fn = os.path.basename(fp)
                    if norm_dir in fp_norm and pat.search(fn):
                        ref_file = fp
                        break
            if ref_file:
                if ref_file not in metadata_cache:
                    metadata_cache[ref_file] = extract_metadata(ref_file)
                meta = metadata_cache[ref_file]
                ws_sc.cell(row=r, column=10).value = (
                    special['force_status'] if special.get('force_status')
                    else meta.status
                )
                ws_sc.cell(row=r, column=11).value = meta.revision
                ws_sc.cell(row=r, column=12).value = meta.date
                for col in [10, 11, 12]:
                    cell = ws_sc.cell(row=r, column=col)
                    if cell.value == 'NA':
                        cell.fill = UNMATCHED_FILL
            else:
                for col in [11, 12]:
                    ws_sc.cell(row=r, column=col).value = 'NA'
                    ws_sc.cell(row=r, column=col).fill = UNMATCHED_FILL
                ws_sc.cell(row=r, column=10).value = (
                    special['force_status'] if special.get('force_status') else 'NA'
                )
                if ws_sc.cell(row=r, column=10).value == 'NA':
                    ws_sc.cell(row=r, column=10).fill = UNMATCHED_FILL
            continue

        # Split keywords by '|' or newline
        keywords = [k.strip() for k in e_val.replace('\n', '|').split('|') if k.strip()]
        is_directory_only = _is_directory_only_entry(d_val)

        all_matched_files = []
        for kw in keywords:
            result = match_keyword_to_file(kw, all_files, min_score)
            if result.is_matched:
                for fp in result.matched_files:
                    if fp not in all_matched_files:
                        all_matched_files.append(fp)

        if all_matched_files:
            matched_count += 1
            first_file = all_matched_files[0]
            if first_file not in metadata_cache:
                metadata_cache[first_file] = extract_metadata(first_file)
            meta = metadata_cache[first_file]

            if is_directory_only:
                dirs = []
                for fp in all_matched_files:
                    d = _get_directory_path(fp, doc_dir)
                    if d not in dirs:
                        dirs.append(d)
                _set_h_cell(ws_sc, r, '\n'.join(dirs), doc_dir, dir_paths=dirs)
            else:
                paths = []
                for fp in all_matched_files:
                    if doc_dir and fp.startswith(doc_dir):
                        paths.append(os.path.relpath(fp, doc_dir))
                    else:
                        paths.append(fp)
                _set_h_cell(ws_sc, r, '\n'.join(paths), doc_dir,
                            file_paths=all_matched_files)

            ws_sc.cell(row=r, column=9).value = meta.author    # I
            ws_sc.cell(row=r, column=10).value = meta.status   # J
            ws_sc.cell(row=r, column=11).value = meta.revision # K
            ws_sc.cell(row=r, column=12).value = meta.date     # L
        else:
            unmatched_count += 1
            for col in [8, 9, 10, 11, 12]:
                cell = ws_sc.cell(row=r, column=col)
                cell.value = 'NA'
                cell.fill = UNMATCHED_FILL

    # --- Incremental diff ---
    if old_safety_case:
        _apply_incremental_diff(ws_sc, old_safety_case)

    # Apply borders to all filled cells in Safety Case
    for r in range(14, ws_sc.max_row + 1):
        for c in [8, 9, 10, 11, 12]:
            ws_sc.cell(row=r, column=c).alignment = CELL_ALIGN
            ws_sc.cell(row=r, column=c).border = THIN_BORDER

    # Save
    wb.save(output_path)
    wb.close()

    # openpyxl drops some drawings during save — inject them back from template
    _restore_drawings(template_path, output_path)

    # Log unmatched + statistics
    print(f"\n[Safety Case Summary]", file=sys.stderr)
    print(f"  Total processed (F=No): {total_processed}", file=sys.stderr)
    print(f"  Matched: {matched_count}", file=sys.stderr)
    print(f"  Unmatched: {unmatched_count}", file=sys.stderr)
    if total_processed > 0:
        print(f"  Match rate: {matched_count/total_processed*100:.1f}%", file=sys.stderr)

    return output_path
