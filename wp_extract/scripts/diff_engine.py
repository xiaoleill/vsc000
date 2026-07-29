"""
Diff Engine Module
Compares old vs. new TR-format report data and produces diff annotations.

Three rules:
1. All cells default to no background (fresh values overwrite old).
2. Previously-matched WP now unmatched → clear metadata cells, log [DROPPED].
3. Value changed from old → fill light-green, log WARNING with old_value.

Row match key: (sheet_name, work_product, doc_no)
"""

import os
import sys
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl
from openpyxl.styles import PatternFill

# Light green for changed cells (Rule 3)
CHANGED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Light grey for dropped rows (Rule 2)
DROPPED_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# Columns to compare in diff (visible data columns, exclude No. which is a sequence number)
COMPARABLE_COLS = [
    'Work Products', 'Document No.', 'Revision', 'Author', 'Reviewer',
    'Approver', 'Status', 'Date', 'Matched File', 'Note',
]

# Columns that get cleared on drop (Rule 2)
CLEAR_ON_DROP = [
    'Document No.', 'Revision', 'Author', 'Reviewer', 'Approver',
    'Status', 'Date', 'Matched File', 'Note',
]

# TR-format column headers (column letter -> header name)
TR_COLUMNS = {
    'A': 'No.', 'B': 'Work Products', 'C': 'Applicable?',
    'D': 'Rational if not applicable', 'E': 'Document No.',
    'F': 'Revision', 'G': 'Author', 'H': 'Reviewer',
    'I': 'Approver', 'J': 'Status', 'K': 'Date', 'L': 'Note',
    'M': 'Matched File',
    'N': 'Phase', 'O': 'Task ID', 'P': 'FS Phase', 'Q': 'Sub Flow',
    'R': 'Plan Start', 'S': 'Plan End', 'T': 'Actual Start', 'U': 'Actual End',
}

# Reverse: header name -> column letter
HEADER_TO_COL = {v: k for k, v in TR_COLUMNS.items()}

# Column letters for comparable columns (excludes No., Applicable?, Rational, hidden cols)
COMPARABLE_COL_LETTERS = {
    col_l for col_l, header in TR_COLUMNS.items()
    if header in COMPARABLE_COLS
}

# Columns that should NOT be cleared when a row is dropped (keep skeleton)
SKELETON_COLS = {'A', 'B', 'C', 'D', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U'}


@dataclass
class DiffCell:
    """A cell with optional diff annotation."""
    value: Any
    is_changed: bool = False
    old_value: str = ""


@dataclass
class DiffRow:
    """A row in the TR-format report with diff info."""
    sheet_name: str
    cells: dict[str, DiffCell] = field(default_factory=dict)  # col_letter -> DiffCell
    is_dropped: bool = False     # Rule 2: old matched → new unmatched
    is_new: bool = False         # New row, not in old report


def _make_key(sheet_name: str, work_product: str, doc_no: str,
              matched_file: str = '') -> tuple:
    """Create a row-matching key. Includes matched_file to distinguish
    duplicate doc_nos from different files (e.g., part5 vs part9)."""
    import os
    # Use basename of matched file as additional discriminator
    file_key = os.path.basename(matched_file).lower() if matched_file else ''
    return (
        sheet_name.strip().lower(),
        (work_product or '').strip().lower(),
        (doc_no or '').strip().lower(),
        file_key,
    )


def _cell_ref(sheet_name: str, col_letter: str, row_num: int) -> str:
    """Format a cell reference like 'PAC_01 (FS-TR1)'!F5."""
    return f"'{sheet_name}'!{col_letter}{row_num}"


def load_old_report(path: str) -> dict:
    """
    Read an existing TR-format report and index its rows.

    Reads all per-phase sheets (those matching 'PAC_*').
    Returns: {key_tuple: {col_letter: (value, row_num)}}
    where key_tuple = (sheet_name, work_product, doc_no)
    """
    if not path or not os.path.exists(path):
        return {}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return {}

    index = {}

    for sn in wb.sheetnames:
        if not sn.startswith('PAC_'):
            continue
        ws = wb[sn]

        # Find header row (row 1, column B = "Work Products")
        header_row = 1
        found_header = False
        for r in range(1, min(5, ws.max_row + 1)):
            b_val = ws.cell(row=r, column=2).value
            if b_val and 'Work Products' in str(b_val):
                header_row = r
                found_header = True
                break
        if not found_header:
            continue

        # Read data rows
        for r in range(header_row + 1, ws.max_row + 1):
            wp = ws.cell(row=r, column=2).value   # B = Work Products
            doc_no = ws.cell(row=r, column=5).value  # E = Document No.

            if not wp or not str(wp).strip():
                continue  # empty row

            wp_str = str(wp).strip()
            doc_str = str(doc_no).strip() if doc_no else ''
            m_val = ws.cell(row=r, column=13).value  # M = Matched File
            mf_str = str(m_val).strip() if m_val else ''
            key = _make_key(sn, wp_str, doc_str, mf_str)

            row_data = {}
            for col_letter in TR_COLUMNS:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                cell = ws.cell(row=r, column=col_idx)
                row_data[col_letter] = (cell.value, r)

            index[key] = row_data

    wb.close()
    return index


def compute_diff(
    new_rows: list[dict],       # list of {col_letter: value} for each new row
    sheet_name: str,
    old_index: dict,            # from load_old_report
    delete_unmatched: bool = False,
) -> tuple[list[DiffRow], list[str]]:
    """
    Compare new rows against old report data for one sheet.

    Args:
        new_rows: List of dicts, each mapping col_letter -> cell value.
        sheet_name: Current sheet name (e.g. 'PAC_01 (FS-TR1)').
        old_index: Full old report index from load_old_report().
        delete_unmatched: If True, deleted rows are marked for removal.

    Returns:
        (list of DiffRow, list of WARNING strings).
    """
    warnings: list[str] = []
    diff_rows: list[DiffRow] = []

    # Track which old keys have been matched by new rows
    matched_old_keys: set[tuple] = set()

    for new_row in new_rows:
        wp = str(new_row.get('B', '') or '').strip()
        doc_no = str(new_row.get('E', '') or '').strip()
        mf = str(new_row.get('M', '') or '').strip()
        key = _make_key(sheet_name, wp, doc_no, mf)

        old_row = old_index.get(key)
        dr = DiffRow(sheet_name=sheet_name)

        if old_row is None:
            # New row — not in old report
            dr.is_new = True
            for col_letter, value in new_row.items():
                dr.cells[col_letter] = DiffCell(value=value)
        else:
            matched_old_keys.add(key)
            old_row_num = old_row['B'][1]  # row number from old report

            # Compare each cell
            for col_letter, new_value in new_row.items():
                old_value, _ = old_row.get(col_letter, (None, None))

                # Normalize for comparison
                old_str = _norm_for_compare(old_value)
                new_str = _norm_for_compare(new_value)

                is_comparable = col_letter in COMPARABLE_COL_LETTERS

                if is_comparable and old_str != new_str:
                    # Rule 3: changed (only for comparable columns)
                    dr.cells[col_letter] = DiffCell(
                        value=new_value, is_changed=True, old_value=old_str
                    )
                    ref = _cell_ref(sheet_name, col_letter, old_row_num)
                    warnings.append(
                        f"WARNING: updated CELL {ref} <{old_str}>"
                    )
                else:
                    # Unchanged or non-comparable (e.g., No. sequence number)
                    dr.cells[col_letter] = DiffCell(value=new_value)

        diff_rows.append(dr)

    # Rule 2: Find old rows that are NOT in new results
    for key, old_row in old_index.items():
        old_sheet, old_wp, old_doc, old_file = key
        if old_sheet != sheet_name.strip().lower():
            continue
        if key in matched_old_keys:
            continue

        # Old row had data but is now unmatched
        old_row_num = old_row['B'][1]
        wp_value = old_row.get('B', ('', 0))[0]
        doc_value = old_row.get('E', ('', 0))[0]
        matched_file = old_row.get('M', ('', 0))[0]

        # Only warn if the old row actually had a matched file (was a real match)
        if matched_file and str(matched_file).strip():
            warnings.append(
                f"WARNING: [DROPPED] Sheet='{sheet_name}', "
                f"WP='{wp_value}', DocNo='{doc_value}' — "
                f"previously matched, now unmatched"
            )

            if not delete_unmatched:
                # Create a dropped row: keep skeleton, clear metadata
                dr = DiffRow(sheet_name=sheet_name, is_dropped=True)
                for col_letter in TR_COLUMNS:
                    if col_letter in SKELETON_COLS:
                        # Keep skeleton columns from old report
                        old_val, _ = old_row.get(col_letter, (None, None))
                        dr.cells[col_letter] = DiffCell(value=old_val)
                    elif col_letter in CLEAR_ON_DROP:
                        old_val, _ = old_row.get(col_letter, (None, None))
                        dr.cells[col_letter] = DiffCell(
                            value=None, is_changed=True,
                            old_value=_norm_for_compare(old_val)
                        )
                    else:
                        old_val, _ = old_row.get(col_letter, (None, None))
                        dr.cells[col_letter] = DiffCell(value=old_val)
                diff_rows.append(dr)
            # If delete_unmatched, don't add the row at all

    return diff_rows, warnings


def _norm_for_compare(value: Any) -> str:
    """Normalize a cell value for comparison."""
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()
