"""
Safety Plan Parser Module
Parses the "Safety Plan (schedule)" sheet from <project>_safety plan.xlsx.

Extracts:
- Development phases (PAC_01 ~ PAC_05) from merged cells in column B
- Work items with Task ID (C), FS Phase (D), Sub Flow (E)
- Input documents (L) and Output WP documents (M)
- Plan/Actual start/end dates (Q, R, S, T)
- Responsibility (N), Judge/Confirm (O), Approval (P)
"""

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.utils import range_boundaries


@dataclass
class WorkItem:
    """Represents a single work item row in the safety plan schedule."""
    phase: str                          # e.g. "PAC_01"
    task_id: str                        # e.g. "FS-1", "FS-TR1"
    fs_phase_name: str                  # e.g. "SEooC definition"
    sub_flow: str                       # e.g. "SEooC definition"
    input_docs: list[str] = field(default_factory=list)   # L column, split by newline
    output_wps: list[str] = field(default_factory=list)   # M column, split by newline
    responsibility: str = ""            # N column
    judge_confirm: str = ""             # O column
    approval: str = ""                  # P column
    plan_start: Optional[datetime] = None   # Q column
    plan_end: Optional[datetime] = None     # R column
    actual_start: Optional[datetime] = None # S column
    actual_end: Optional[datetime] = None   # T column
    row_num: int = 0                    # original Excel row number


@dataclass
class PhaseInfo:
    """Development phase information."""
    name: str                           # e.g. "PAC_01"
    start_row: int
    end_row: int
    work_items: list[WorkItem] = field(default_factory=list)


def _get_merged_cell_map(ws, col_letter: str) -> dict[int, str]:
    """
    Build a map: row_number -> merged cell value for a given column.
    Handles merged cells by expanding the value to all covered rows.
    """
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    merged_map = {}

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_col <= col_idx <= max_col:
            value = ws.cell(row=min_row, column=col_idx).value
            for r in range(min_row, max_row + 1):
                merged_map[r] = value

    return merged_map


def _split_cell_value(value) -> list[str]:
    """Split a cell value by newline, filter empty strings.
    Merges standalone parenthetical qualifiers (e.g., "(specification phase)")
    with the preceding keyword.
    """
    if value is None:
        return []
    if isinstance(value, datetime):
        return []
    text = str(value).strip()
    if not text or text == '-' or text == 'n/a':
        return []

    raw_items = [s.strip().rstrip('"').rstrip("'").rstrip('"').rstrip('”').rstrip('“')
                 for s in text.split('\n') if s.strip() and s.strip() != '-']

    # Merge standalone parenthetical lines with the previous item
    # e.g., "Dependent failure analysis (DFA)" + "(specification phase)"
    #   -> "Dependent failure analysis (DFA) (specification phase)"
    merged = []
    for item in raw_items:
        # Check if this is a standalone parenthetical qualifier
        # e.g., "(specification phase)", "(refined)", "(post layout phase)"
        is_standalone_paren = (
            (item.startswith('(') and item.endswith(')')) or
            (item.startswith('（') and item.endswith('）'))
        )
        if is_standalone_paren and merged:
            # Append to previous item
            merged[-1] = merged[-1] + ' ' + item
        else:
            merged.append(item)

    return merged


def _parse_date(value) -> Optional[datetime]:
    """Parse a cell value into a datetime, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text or text == '-' or text.lower() == 'n/a':
        return None
    # Try common date formats
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _normalize_text(text: str) -> str:
    """Normalize whitespace in text."""
    if not text:
        return ""
    return ' '.join(text.split())


def parse_safety_plan(filepath: str) -> list[WorkItem]:
    """
    Parse the safety plan Excel file and return all work items.

    Args:
        filepath: Path to the <project>_safety plan.xlsx file.

    Returns:
        List of WorkItem objects, starting from PAC_01 (Charter phase excluded).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_name = 'Safety Plan (schedule)'
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # Build merged cell maps for columns B and C
    b_merged = _get_merged_cell_map(ws, 'B')
    c_merged = _get_merged_cell_map(ws, 'C')
    d_merged = _get_merged_cell_map(ws, 'D')

    # Find the row where Charter data begins (row with B="Charter")
    charter_row = None
    for r in range(1, ws.max_row + 1):
        b_val = b_merged.get(r) or ws.cell(row=r, column=2).value
        if b_val and str(b_val).strip() == 'Charter':
            charter_row = r
            break

    if charter_row is None:
        raise ValueError("Could not find 'Charter' phase in column B")

    # Determine the start row: find the first PAC_xx row after Charter
    start_row = None
    for r in range(charter_row + 1, ws.max_row + 1):
        b_val = b_merged.get(r) or ws.cell(row=r, column=2).value
        if b_val and str(b_val).strip().startswith('PAC_'):
            start_row = r
            break

    if start_row is None:
        raise ValueError("Could not find any PAC_xx phase after Charter")

    work_items = []

    # Iterate from start_row to end
    current_phase = ""
    current_task_id = ""
    current_fs_phase = ""

    for r in range(start_row, ws.max_row + 1):
        # Get phase from B column (merged map)
        b_val = b_merged.get(r) or ws.cell(row=r, column=2).value
        if b_val:
            b_text = str(b_val).strip()
            if b_text.startswith('PAC_'):
                current_phase = b_text
            # Skip header rows (where B is the phase but C/D are header-like)
            # Header rows have C values like "Planning", "Specification", etc.
            c_val = ws.cell(row=r, column=3).value
            if c_val:
                c_text = str(c_val).strip()
                # Check if this is a phase header row (not a real work item)
                if c_text in ('Planning', 'Specification', 'Design', 'Evaluation',
                              'Production') or not c_text.startswith('FS-'):
                    continue

        # Get task ID from C column (merged map + direct)
        c_val = c_merged.get(r) or ws.cell(row=r, column=3).value
        if c_val:
            c_text = str(c_val).strip()
            if c_text.startswith('FS-'):
                current_task_id = c_text
                # Also update FS phase from D column
                d_val = d_merged.get(r) or ws.cell(row=r, column=4).value
                if d_val:
                    current_fs_phase = _normalize_text(str(d_val))
            else:
                continue  # skip non-FS rows (headers)

        # Get values from other columns
        e_val = ws.cell(row=r, column=5).value
        l_val = ws.cell(row=r, column=12).value
        m_val = ws.cell(row=r, column=13).value
        n_val = ws.cell(row=r, column=14).value
        o_val = ws.cell(row=r, column=15).value
        p_val = ws.cell(row=r, column=16).value

        sub_flow = _normalize_text(str(e_val)) if e_val else ""

        # Skip rows that are clearly sub-items with no real output WP
        # (e.g., rows that only have an IP name in L column and very specific M value)
        # We keep all rows but flag empty output WPs

        input_docs = _split_cell_value(l_val)
        output_wps = _split_cell_value(m_val)

        # Skip rows with no output WPs and no sub_flow
        if not output_wps and not sub_flow:
            continue

        wi = WorkItem(
            phase=current_phase,
            task_id=current_task_id,
            fs_phase_name=current_fs_phase,
            sub_flow=sub_flow,
            input_docs=input_docs,
            output_wps=output_wps,
            responsibility=_normalize_text(str(n_val)) if n_val else "",
            judge_confirm=_normalize_text(str(o_val)) if o_val else "",
            approval=_normalize_text(str(p_val)) if p_val else "",
            plan_start=_parse_date(ws.cell(row=r, column=17).value),
            plan_end=_parse_date(ws.cell(row=r, column=18).value),
            actual_start=_parse_date(ws.cell(row=r, column=19).value),
            actual_end=_parse_date(ws.cell(row=r, column=20).value),
            row_num=r,
        )
        work_items.append(wi)

    wb.close()
    return work_items


def get_all_output_keywords(work_items: list[WorkItem]) -> list[str]:
    """
    Extract all unique output WP keywords from work items.
    Deduplicated, preserving order.
    """
    seen = set()
    result = []
    for wi in work_items:
        for kw in wi.output_wps:
            if kw.lower() not in seen:
                seen.add(kw.lower())
                result.append(kw)
    return result
