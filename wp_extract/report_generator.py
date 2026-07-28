"""
Report Generator Module — TR Format
Generates an Excel report matching the "Subject to Review" layout from
the Functional Safety Technical Review checklist.

Output: per-phase sheets (PAC_01~PAC_05) + Statistics.
Supports incremental diff mode against a previous report.
"""

import os
import sys
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from safety_plan_parser import WorkItem
from file_matcher import MatchResult
from metadata_extractor import DocumentMetadata, NOT_FOUND, extract_metadata

# Phase order and sheet naming
PHASE_SHEET_MAP = {
    'PAC_01': 'PAC_01 (FS-TR1)',
    'PAC_02': 'PAC_02 (FS-TR2)',
    'PAC_03': 'PAC_03 (FS-TR3)',
    'PAC_04': 'PAC_04 (FS-TR4)',
    'PAC_05': 'PAC_05 (FS-TR5)',
}

# TR-format column definitions
# (col_letter, header, width, hidden)
TR_COLS = [
    ('A', 'No.',              6,  False),
    ('B', 'Work Products',    45, False),
    ('C', 'Applicable?',      12, False),
    ('D', 'Rational if not applicable', 30, False),
    ('E', 'Document No.',     28, False),
    ('F', 'Revision',         10, False),
    ('G', 'Author',           18, False),
    ('H', 'Reviewer',         20, False),
    ('I', 'Approver',         15, False),
    ('J', 'Status',           12, False),
    ('K', 'Date',             14, False),
    ('L', 'Note',             30, False),
    ('M', 'Matched File',     55, False),
    # Hidden columns
    ('N', 'Phase',            10, True),
    ('O', 'Task ID',          10, True),
    ('P', 'FS Phase',         28, True),
    ('Q', 'Sub Flow',         28, True),
    ('R', 'Plan Start',       14, True),
    ('S', 'Plan End',         14, True),
    ('T', 'Actual Start',     14, True),
    ('U', 'Actual End',       14, True),
]

# Column letter -> header name
COL_HEADERS = {c: h for c, h, _, _ in TR_COLS}
# Column letter -> width
COL_WIDTHS = {c: w for c, _, w, _ in TR_COLS}
# Hidden columns
HIDDEN_COLS = {c for c, _, _, h in TR_COLS if h}

# Styles
HEADER_FONT = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
# Diff colors (from diff_engine)
CHANGED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
NOT_FOUND_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # yellow for unmatched rows
NA_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')          # light red for NA cells

# Descriptive-note indicator phrases
DESCRIPTIVE_PATTERNS = [
    'refer ', 'same with', 'not a', 'n/a', 'optional', 'the wps',
    'all work products', 'internal target specification',
]


def _is_descriptive_note(keyword: str) -> bool:
    """Heuristic: is this keyword a descriptive note rather than a real document?"""
    kw = keyword.strip().lower()
    for pat in DESCRIPTIVE_PATTERNS:
        if pat in kw:
            return True
    # Starts with lowercase letter (not a proper document name)
    if kw and kw[0].islower():
        return True
    return False


def _fmt_date(dt) -> str:
    """Format a date for display as YYYY-MM-DD."""
    import re as _re
    if dt is None:
        return ''
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d')
    text = str(dt).strip()
    if not text:
        return ''
    # Normalize YYYY/M/D or YYYY.M.D to YYYY-MM-DD
    m = _re.match(r'(\d{4})[/\-\.](\d{1,2})[/\-\.](\d{1,2})', text)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    return text


def _determine_applicable(matched_file: Optional[str], meta: Optional[DocumentMetadata],
                          keyword: str) -> tuple[str, str]:
    """
    Determine Applicable? and Rational values.
    Returns (applicable, rational).
    """
    if _is_descriptive_note(keyword):
        return 'No', 'Descriptive note in Safety Plan'

    if matched_file is None:
        return 'Check', 'File not found in doc-dir'

    if meta:
        status = (meta.status or '').strip().lower()
        if 'released' in status:
            return 'Yes', '-'
        elif status and status != NOT_FOUND.lower():
            return 'Yes', f'Status: {meta.status}'
        else:
            return 'Yes', '-'

    return 'Yes', '-'


def _build_note(matched_file: Optional[str], keyword: str, is_multi: bool,
                score: float = 0.0, ext_mismatch: bool = False,
                ext_expected: str = '') -> str:
    """Build the Note column content."""
    notes = []
    if matched_file is None:
        if _is_descriptive_note(keyword):
            return ''
        return 'File not found in doc-dir'
    if is_multi and score > 0:
        notes.append(f'Multi-doc keyword (score: {score:.0%})')
    if ext_mismatch and ext_expected:
        actual_ext = os.path.splitext(matched_file)[1]
        notes.append(f'Format mismatch: expected {ext_expected}, got {actual_ext}')
    return '; '.join(notes) if notes else ''


def generate_tr_report(
    work_items: list[WorkItem],
    match_results: dict[str, MatchResult],
    output_path: str,
    doc_dir: str = "",
    old_report_path: str = None,
    delete_unmatched: bool = False,
) -> str:
    """
    Generate TR-format Excel report with per-phase sheets.

    Args:
        work_items: All work items from safety plan.
        match_results: Dict mapping keyword -> MatchResult.
        output_path: Output Excel file path.
        doc_dir: Document directory (for relative paths).
        old_report_path: Path to previous report for diff (None = fresh mode).
        delete_unmatched: If True, delete entire dropped rows.

    Returns:
        Path to the generated report.
    """
    # --- Diff engine (optional) ---
    old_index = {}
    diff_warnings: list[str] = []
    do_diff = old_report_path and os.path.exists(old_report_path)
    if do_diff:
        from diff_engine import load_old_report, compute_diff
        old_index = load_old_report(old_report_path)

    # --- Build keyword -> MatchResult lookup ---
    kw_to_result: dict[str, MatchResult] = {}
    for result in match_results.values():
        kw_to_result[result.keyword] = result

    # --- Metadata cache ---
    metadata_cache: dict[str, DocumentMetadata] = {}

    # --- Group work items by phase ---
    phase_items: dict[str, list[WorkItem]] = {}
    for wi in work_items:
        phase_items.setdefault(wi.phase, []).append(wi)

    # --- Create workbook ---
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    total_matched_count = 0
    total_unmatched_count = 0
    total_file_count = 0

    # --- Generate per-phase sheets ---
    for phase in ['PAC_01', 'PAC_02', 'PAC_03', 'PAC_04', 'PAC_05']:
        sheet_name = PHASE_SHEET_MAP.get(phase, phase)
        ws = wb.create_sheet(sheet_name)
        items = phase_items.get(phase, [])

        # --- Build new row data for this sheet ---
        new_rows: list[dict] = []  # list of {col_letter: value}
        row_num = 1

        for wi in items:
            for kw in wi.output_wps:
                result = kw_to_result.get(kw)
                is_matched = result and result.is_matched

                if is_matched and result.matched_files:
                    for fi, filepath in enumerate(result.matched_files):
                        score = result.scores[fi] if fi < len(result.scores) else 0
                        is_multi = len(result.matched_files) > 1

                        # Metadata
                        if filepath not in metadata_cache:
                            metadata_cache[filepath] = extract_metadata(filepath)
                        meta = metadata_cache[filepath]

                        display_path = filepath
                        if doc_dir and filepath.startswith(doc_dir):
                            display_path = os.path.relpath(filepath, doc_dir)

                        applicable, rational = _determine_applicable(filepath, meta, kw)
                        note = _build_note(filepath, kw, is_multi, score,
                                          result.extension_mismatch,
                                          result.extension_expected)

                        row_data = {
                            'A': None,  # No. filled later
                            'B': kw,
                            'C': applicable,
                            'D': rational,
                            'E': meta.document_no if meta else NOT_FOUND,
                            'F': meta.revision if meta else NOT_FOUND,
                            'G': meta.author if meta else NOT_FOUND,
                            'H': meta.reviewer if meta else NOT_FOUND,
                            'I': meta.approver if meta else NOT_FOUND,
                            'J': meta.status if meta else NOT_FOUND,
                            'K': meta.date if meta else NOT_FOUND,
                            'L': note,
                            'M': display_path,
                            'N': wi.phase,
                            'O': wi.task_id,
                            'P': wi.fs_phase_name,
                            'Q': wi.sub_flow,
                            'R': _fmt_date(wi.plan_start),
                            'S': _fmt_date(wi.plan_end),
                            'T': _fmt_date(wi.actual_start),
                            'U': _fmt_date(wi.actual_end),
                        }
                        new_rows.append(row_data)
                        total_file_count += 1
                else:
                    # Unmatched keyword
                    applicable, rational = _determine_applicable(None, None, kw)
                    note = _build_note(None, kw, False)

                    row_data = {
                        'A': None,
                        'B': kw,
                        'C': applicable,
                        'D': rational,
                        'E': NOT_FOUND,
                        'F': NOT_FOUND,
                        'G': NOT_FOUND,
                        'H': NOT_FOUND,
                        'I': NOT_FOUND,
                        'J': NOT_FOUND,
                        'K': NOT_FOUND,
                        'L': note,
                        'M': '',
                        'N': wi.phase,
                        'O': wi.task_id,
                        'P': wi.fs_phase_name,
                        'Q': wi.sub_flow,
                        'R': _fmt_date(wi.plan_start),
                        'S': _fmt_date(wi.plan_end),
                        'T': _fmt_date(wi.actual_start),
                        'U': _fmt_date(wi.actual_end),
                    }
                    new_rows.append(row_data)
                    total_unmatched_count += 1

        total_matched_count += sum(
            1 for r in new_rows if r.get('M') and r['M'] != ''
        )

        # --- Compute diff if incremental mode ---
        diff_rows = None
        if do_diff and old_index:
            from diff_engine import compute_diff, CHANGED_FILL as CF
            diff_rows, warnings = compute_diff(
                new_rows, sheet_name, old_index, delete_unmatched
            )
            diff_warnings.extend(warnings)

        # --- Write sheet ---
        # Header row
        for col_letter, header, width, hidden in TR_COLS:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

        # Set column widths and hidden
        for col_letter, _, width, hidden in TR_COLS:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            ws.column_dimensions[col_letter].width = width
            if hidden:
                ws.column_dimensions[col_letter].hidden = True

        # Freeze header
        ws.freeze_panes = 'A2'

        # Write data rows
        source_rows = diff_rows if diff_rows is not None else None
        if source_rows is not None:
            # Incremental mode with diff
            for r_idx, dr in enumerate(source_rows, 2):
                seq = r_idx - 1
                for col_letter, _, _, _ in TR_COLS:
                    col_idx = openpyxl.utils.column_index_from_string(col_letter)
                    dc = dr.cells.get(col_letter)
                    value = dc.value if dc else None

                    # Sequence number for column A
                    if col_letter == 'A':
                        value = seq

                    cell = ws.cell(row=r_idx, column=col_idx, value=value)
                    cell.alignment = CELL_ALIGN
                    cell.border = THIN_BORDER

                    # Apply diff coloring (Rule 3: changed → light green)
                    if dc and dc.is_changed:
                        cell.fill = CHANGED_FILL
                    elif isinstance(value, str) and value == 'NA':
                        cell.fill = NA_FILL
        else:
            # Fresh mode
            for r_idx, row_data in enumerate(new_rows, 2):
                seq = r_idx - 1
                for col_letter, _, _, _ in TR_COLS:
                    col_idx = openpyxl.utils.column_index_from_string(col_letter)
                    value = row_data.get(col_letter)
                    if col_letter == 'A':
                        value = seq

                    cell = ws.cell(row=r_idx, column=col_idx, value=value)
                    cell.alignment = CELL_ALIGN
                    cell.border = THIN_BORDER
                    if isinstance(value, str) and value == 'NA':
                        cell.fill = NA_FILL

    # --- Statistics sheet ---
    ws_stats = wb.create_sheet("Statistics")
    total_kw = sum(len(wi.output_wps) for wi in work_items)

    stats = [
        ['Metric', 'Value'],
        ['Total Work Items', len(work_items)],
        ['Total Output WP Keywords', total_kw],
        ['Matched File Entries', total_file_count],
        ['Unmatched Keywords', total_unmatched_count],
        ['Document Directory', doc_dir],
        ['Mode', 'Incremental' if do_diff else 'Fresh'],
        ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]

    for i, (label, value) in enumerate(stats, 1):
        ca = ws_stats.cell(row=i, column=1, value=label)
        cb = ws_stats.cell(row=i, column=2, value=value)
        if i == 1:
            ca.font = HEADER_FONT
            ca.fill = HEADER_FILL
            cb.font = HEADER_FONT
            cb.fill = HEADER_FILL
        ca.alignment = CELL_ALIGN
        cb.alignment = CELL_ALIGN
        ca.border = THIN_BORDER
        cb.border = THIN_BORDER

    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 50

    # --- Save ---
    wb.save(output_path)

    # --- Print diff warnings to stderr ---
    if diff_warnings:
        print("\n[D I F F   W A R N I N G S]", file=sys.stderr)
        print("-" * 60, file=sys.stderr)
        for w in diff_warnings:
            print(w, file=sys.stderr)
        print("-" * 60, file=sys.stderr)
        print(f"Total warnings: {len(diff_warnings)}\n", file=sys.stderr)

    return output_path
