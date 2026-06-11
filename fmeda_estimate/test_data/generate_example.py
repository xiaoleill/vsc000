"""
generate_example.py — 生成测试用Excel示例文件

运行此脚本生成 example_input.xlsx，包含3个sheet：
  - Info: 芯片基本信息（含ASIL等级）
  - Module: 模块面积占比
  - FMEDA: 失效模式、DC、DC_LF（含合并单元格）
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def apply_header_style(ws, row, col_count):
    """应用表头样式"""
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def apply_cell_style(ws, row, col_count):
    """应用数据单元格样式"""
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = cell_align
        cell.border = thin_border


def main():
    wb = Workbook()

    # =========================================================================
    # Sheet 1: Info
    # =========================================================================
    ws_info = wb.active
    ws_info.title = "Info"

    info_data = [
        ["参数", "值"],
        ["项目名称(project)", "SAFETY_CHIP_X1"],
        ["芯片总面积(chip_area)", 99999.00],
        ["总体失效率(λ_chip/FIT)", 100.00],
        ["模块数目(M_num)", 3],
        ["ASIL等级", "B"],
    ]

    for row_data in info_data:
        ws_info.append(row_data)

    apply_header_style(ws_info, 1, 2)
    for r in range(2, len(info_data) + 1):
        apply_cell_style(ws_info, r, 2)
    ws_info.column_dimensions['A'].width = 28
    ws_info.column_dimensions['B'].width = 20

    # =========================================================================
    # Sheet 2: Module
    # =========================================================================
    ws_module = wb.create_sheet("Module")

    module_data = [
        ["模块(M)", "占比(MD%)"],
        ["M0_CPU_Core", 0.40],
        ["M1_Memory", 0.35],
        ["M2_Peripheral", 0.25],
    ]

    for row_data in module_data:
        ws_module.append(row_data)

    apply_header_style(ws_module, 1, 2)
    for r in range(2, len(module_data) + 1):
        apply_cell_style(ws_module, r, 2)
    ws_module.column_dimensions['A'].width = 20
    ws_module.column_dimensions['B'].width = 18

    # =========================================================================
    # Sheet 3: FMEDA（含DC_LF列）
    # =========================================================================
    ws_fmeda = wb.create_sheet("FMEDA")

    fmeda_headers = [
        "模块(M)", "失效模式(FM)", "失效占比(FMD%)",
        "诊断覆盖率(DC%)", "潜在故障DC(DC_LF%)"
    ]
    ws_fmeda.append(fmeda_headers)

    # 测试数据：3个模块 × 2~3个失效模式
    # DC_LF通常比DC低，因为潜在故障更难检测
    fmeda_rows = [
        # module, fm, fmd%, dc%, dc_lf%
        ("M0_CPU_Core", "FM0_ALU_Parity_Error",     0.35, 0.95, 0.90),
        ("M0_CPU_Core", "FM1_Register_Soft_Error",   0.40, 0.90, 0.85),
        ("M0_CPU_Core", "FM2_Control_Unit_SEU",      0.25, 0.85, 0.80),
        ("M1_Memory",   "FM0_SRAM_Multi_Bit_Upset",  0.55, 0.92, 0.88),
        ("M1_Memory",   "FM1_ECC_Decoder_Failure",   0.45, 0.88, 0.82),
        ("M2_Peripheral","FM0_SPI_Timeout_Error",    0.60, 0.80, 0.75),
        ("M2_Peripheral","FM1_GPIO_Stuck_At_Fault",  0.40, 0.75, 0.70),
    ]

    for row_data in fmeda_rows:
        ws_fmeda.append(list(row_data))

    # 合并单元格：模块名称列
    ws_fmeda.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    ws_fmeda.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
    ws_fmeda.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)

    apply_header_style(ws_fmeda, 1, 5)
    for r in range(2, len(fmeda_rows) + 2):
        apply_cell_style(ws_fmeda, r, 5)
    ws_fmeda.column_dimensions['A'].width = 20
    ws_fmeda.column_dimensions['B'].width = 30
    ws_fmeda.column_dimensions['C'].width = 20
    ws_fmeda.column_dimensions['D'].width = 20
    ws_fmeda.column_dimensions['E'].width = 22

    # =========================================================================
    # 保存
    # =========================================================================
    output_path = os.path.join(os.path.dirname(__file__), "example_input.xlsx")
    wb.save(output_path)
    print(f"示例Excel文件已生成：{output_path}")


if __name__ == "__main__":
    main()
