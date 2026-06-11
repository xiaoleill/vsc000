"""
excel_io.py — Excel读写服务

使用openpyxl实现FMEDA评估数据的导入和导出。

导入：
  - Sheet 'Info': 芯片基本信息（含ASIL等级）
  - Sheet 'Module': 模块名称及面积占比
  - Sheet 'FMEDA': 模块失效模式、占比、DC、DC_LF（含合并单元格处理）

导出：
  - 保留原有Info/Module/FMEDA三个sheet（含更新后的数据）
  - 新增Sheet 'Result'：SPFM和LFM计算结果及ASIL合规判断
"""
import os
from typing import List, Tuple, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import (
    SHEET_INFO, SHEET_MODULE, SHEET_FMEDA, SHEET_RESULT,
    MODULE_HEADERS, FMEDA_HEADERS, DEFAULT_ASIL
)
from models.chip_info import ChipInfo
from models.module import Module
from models.fmeda import FailureMode


class ExcelIO:
    """Excel读写服务类"""

    # =========================================================================
    # 样式定义
    # =========================================================================
    HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def _apply_header_style(ws, row: int, col_count: int):
        """为表头行应用统一样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = ExcelIO.HEADER_FONT
            cell.fill = ExcelIO.HEADER_FILL
            cell.alignment = ExcelIO.HEADER_ALIGNMENT
            cell.border = ExcelIO.THIN_BORDER

    @staticmethod
    def _apply_cell_style(ws, row: int, col_count: int):
        """为数据行应用统一样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = ExcelIO.CELL_ALIGNMENT
            cell.border = ExcelIO.THIN_BORDER

    # =========================================================================
    # 导入逻辑
    # =========================================================================

    @staticmethod
    def import_excel(filepath: str) -> Tuple[ChipInfo, List[Module], List[FailureMode]]:
        """从Excel文件导入FMEDA评估所需的全部数据"""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"找不到文件: {filepath}")

        wb = load_workbook(filepath, data_only=True)

        chip_info = ExcelIO._read_info_sheet(wb)
        modules = ExcelIO._read_module_sheet(wb)
        fmeda_list = ExcelIO._read_fmeda_sheet(wb)

        wb.close()
        return chip_info, modules, fmeda_list

    @staticmethod
    def _read_info_sheet(wb: Workbook) -> ChipInfo:
        """读取Info sheet，提取芯片基本信息（含ASIL等级）"""
        if SHEET_INFO not in wb.sheetnames:
            raise ValueError(f"Excel中缺少 '{SHEET_INFO}' sheet")

        ws = wb[SHEET_INFO]
        chip_info = ChipInfo()

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2,
                                values_only=True):
            if row[0] is None:
                continue
            param_name = str(row[0]).strip()
            value = row[1]

            if "project" in param_name.lower() or "项目" in param_name or "名称" in param_name:
                chip_info.project = str(value) if value is not None else ChipInfo.project
            elif "面积" in param_name or "chip_area" in param_name.lower():
                chip_info.chip_area = float(value) if value is not None else ChipInfo.chip_area
            elif "失效率" in param_name or "λ" in param_name or "lambda" in param_name.lower():
                chip_info.lambda_chip = float(value) if value is not None else ChipInfo.lambda_chip
            elif "模块数" in param_name or "M_num" in param_name or "m_num" in param_name.lower():
                chip_info.m_num = int(value) if value is not None else ChipInfo.m_num
            elif "asil" in param_name.lower() or "ASIL" in param_name:
                asil_val = str(value).strip().upper() if value is not None else DEFAULT_ASIL
                if asil_val in ('A', 'B', 'C', 'D'):
                    chip_info.asil = asil_val

        return chip_info

    @staticmethod
    def _read_module_sheet(wb: Workbook) -> List[Module]:
        """读取Module sheet"""
        if SHEET_MODULE not in wb.sheetnames:
            raise ValueError(f"Excel中缺少 '{SHEET_MODULE}' sheet")

        ws = wb[SHEET_MODULE]
        modules = []

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2,
                                values_only=True):
            if row[0] is None:
                continue
            name = str(row[0]).strip()
            if not name:
                continue
            area_pct = float(row[1]) if row[1] is not None else 0.0
            modules.append(Module(name=name, area_pct=area_pct))

        return modules

    @staticmethod
    def _read_fmeda_sheet(wb: Workbook) -> List[FailureMode]:
        """读取FMEDA sheet，支持5列格式（含DC_LF列）或4列格式（兼容旧版）

        新格式（5列）：模块(M) | 失效模式(FM) | 失效占比(FMD%) | 诊断覆盖率(DC%) | 潜在故障DC(DC_LF%)
        旧格式（4列）：模块(M) | 失效模式(FM) | 失效占比(FMD%) | 诊断覆盖率(DC%)
                      此时 DC_LF 默认等于 DC
        """
        if SHEET_FMEDA not in wb.sheetnames:
            raise ValueError(f"Excel中缺少 '{SHEET_FMEDA}' sheet")

        ws = wb[SHEET_FMEDA]
        fmeda_list = []

        # 判断列数（通过读取表头）
        max_col = ws.max_column
        has_dc_lf = (max_col >= 5)

        # 构建合并单元格映射
        merge_map = {}
        for merge_range in ws.merged_cells.ranges:
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    merge_map[(row, col)] = (merge_range.min_row, merge_range.min_col)

        for row_idx in range(2, ws.max_row + 1):
            # 获取模块名称（处理合并单元格）
            if (row_idx, 1) in merge_map:
                top_r, top_c = merge_map[(row_idx, 1)]
                module_name = str(ws.cell(row=top_r, column=top_c).value or "").strip()
            else:
                module_name = str(ws.cell(row=row_idx, column=1).value or "").strip()

            if not module_name:
                continue

            fm_name = str(ws.cell(row=row_idx, column=2).value or "").strip()
            if not fm_name:
                continue

            fm_pct_val = ws.cell(row=row_idx, column=3).value
            fm_pct = float(fm_pct_val) if fm_pct_val is not None else 0.0

            dc_val = ws.cell(row=row_idx, column=4).value
            dc = float(dc_val) if dc_val is not None else 0.0

            # DC_LF：新格式读取第5列，旧格式默认等于DC
            if has_dc_lf:
                dc_lf_val = ws.cell(row=row_idx, column=5).value
                dc_lf = float(dc_lf_val) if dc_lf_val is not None else dc
            else:
                dc_lf = dc

            fmeda_list.append(FailureMode(
                module_name=module_name,
                fm_name=fm_name,
                fm_pct=fm_pct,
                dc=dc,
                dc_lf=dc_lf,
            ))

        return fmeda_list

    # =========================================================================
    # 导出逻辑
    # =========================================================================

    @staticmethod
    def export_excel(
        filepath: str,
        chip_info: ChipInfo,
        modules: List[Module],
        fmeda_list: List[FailureMode],
        spfm: float,
        lfm: float = 0.0,
    ):
        """将当前数据及计算结果导出到Excel文件"""
        wb = Workbook()

        # ---- Sheet 1: Info ----
        ws_info = wb.active
        ws_info.title = SHEET_INFO
        ws_info.append(["参数", "值"])
        for param, val in chip_info.to_dict().items():
            ws_info.append([param, val])
        ExcelIO._apply_header_style(ws_info, 1, 2)
        for r in range(2, ws_info.max_row + 1):
            ExcelIO._apply_cell_style(ws_info, r, 2)
        ws_info.column_dimensions['A'].width = 28
        ws_info.column_dimensions['B'].width = 20

        # ---- Sheet 2: Module ----
        ws_module = wb.create_sheet(SHEET_MODULE)
        export_headers = MODULE_HEADERS + ["失效率(λ_M/FIT)"]
        ws_module.append(export_headers)
        for m in modules:
            ws_module.append([
                m.name,
                round(m.area_pct, 4),
                round(m.lambda_m, 2),
            ])
        ExcelIO._apply_header_style(ws_module, 1, len(export_headers))
        for r in range(2, ws_module.max_row + 1):
            ExcelIO._apply_cell_style(ws_module, r, len(export_headers))
        ws_module.column_dimensions['A'].width = 20
        ws_module.column_dimensions['B'].width = 18
        ws_module.column_dimensions['C'].width = 22

        # ---- Sheet 3: FMEDA ----
        ws_fmeda = wb.create_sheet(SHEET_FMEDA)
        export_fmeda_headers = FMEDA_HEADERS + ["残余失效(λ_R/FIT)", "潜在失效(λ_L/FIT)"]
        ws_fmeda.append(export_fmeda_headers)

        current_row = 2
        fmeda_by_module = {}
        for fm in fmeda_list:
            fmeda_by_module.setdefault(fm.module_name, []).append(fm)

        for module_name, fms in fmeda_by_module.items():
            start_row = current_row
            for i, fm in enumerate(fms):
                ws_fmeda.append([
                    module_name if i == 0 else "",
                    fm.fm_name,
                    round(fm.fm_pct, 4),
                    round(fm.dc, 4),
                    round(fm.dc_lf, 4),
                    round(fm.lambda_r, 2),
                    round(fm.lambda_latent, 2),
                ])
                current_row += 1
            end_row = current_row - 1
            if len(fms) > 1:
                ws_fmeda.merge_cells(
                    start_row=start_row, start_column=1,
                    end_row=end_row, end_column=1
                )

        ExcelIO._apply_header_style(ws_fmeda, 1, len(export_fmeda_headers))
        for r in range(2, ws_fmeda.max_row + 1):
            ExcelIO._apply_cell_style(ws_fmeda, r, len(export_fmeda_headers))
        ws_fmeda.column_dimensions['A'].width = 18
        ws_fmeda.column_dimensions['B'].width = 24
        ws_fmeda.column_dimensions['C'].width = 20
        ws_fmeda.column_dimensions['D'].width = 20
        ws_fmeda.column_dimensions['E'].width = 22
        ws_fmeda.column_dimensions['F'].width = 22
        ws_fmeda.column_dimensions['G'].width = 22

        # ---- Sheet 4: Result ----
        from services.calculator import Calculator
        compliance = Calculator.check_asil_compliance(spfm, lfm, chip_info.asil)

        ws_result = wb.create_sheet(SHEET_RESULT)
        ws_result.append(["指标", "计算结果", "ASIL阈值", "判定", "说明"])

        for metric_name in ['SPFM', 'LFM']:
            info = compliance[metric_name]
            threshold_str = f"≥{info['threshold']*100:.0f}%" if info['threshold'] is not None else "无要求"
            verdict = "PASS" if info['pass'] is True else ("FAIL" if info['pass'] is False else "N/A")
            ws_result.append([
                metric_name,
                f"{info['value']*100:.2f}%",
                threshold_str,
                verdict,
                info['label'],
            ])

        # 汇总
        total_lambda_r = sum(fm.lambda_r for fm in fmeda_list)
        total_lambda_l = sum(fm.lambda_latent for fm in fmeda_list)
        ws_result.append([])
        ws_result.append(["汇总", "", "", "", ""])
        ws_result.append(["λ_R_sum (总残余失效)", f"{total_lambda_r:.2f} FIT", "", "", ""])
        ws_result.append(["λ_L_sum (总潜在失效)", f"{total_lambda_l:.2f} FIT", "", "", ""])
        ws_result.append(["λ_chip", f"{chip_info.lambda_chip:.2f} FIT", "", "", ""])
        ws_result.append(["ASIL等级", chip_info.asil, "", "", ""])

        ExcelIO._apply_header_style(ws_result, 1, 5)
        for r in range(2, ws_result.max_row + 1):
            ExcelIO._apply_cell_style(ws_result, r, 5)
        ws_result.column_dimensions['A'].width = 26
        ws_result.column_dimensions['B'].width = 18
        ws_result.column_dimensions['C'].width = 16
        ws_result.column_dimensions['D'].width = 10
        ws_result.column_dimensions['E'].width = 40

        # 先保存到临时文件，再原子替换，避免Windows下覆盖已有文件时的
        # PermissionError [Errno 13] 问题（文件可能被资源管理器等进程锁定）
        import tempfile
        tmp_fd, tmp_path = tempfile.mkstemp(
            suffix=".xlsx", prefix="fmeda_",
            dir=os.path.dirname(filepath) or os.path.dirname(os.path.abspath(filepath))
        )
        os.close(tmp_fd)
        try:
            wb.save(tmp_path)
            # 删除原有文件（如果存在）
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except PermissionError:
                    # 删除失败时尝试重命名旧文件作为备份
                    backup = filepath + ".bak"
                    if os.path.exists(backup):
                        os.remove(backup)
                    os.rename(filepath, backup)
            os.replace(tmp_path, filepath)
        finally:
            # 清理临时文件（如果还存在）
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
